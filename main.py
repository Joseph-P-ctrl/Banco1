from flask import Flask, render_template, request, send_file, session, redirect, url_for
from AccountService import AccountService
from InterbankService import InterbankService
from ProviderService import ProviderService
from TransferService import TransferService
from BaseDatosService import BaseDatosService
from AsientoService import AsientoService
from io import BytesIO
from flask_session import Session
from flask_caching import Cache
import pandas as pd
import os
import openpyxl
import logging
import traceback
import json
from dotenv import load_dotenv
from functools import wraps
from storage_paths import ensure_data_dirs, bootstrap_bd_from_source, files_path, logs_path, SESSION_DIR
import CorreoService as correo_service

# setup logging
ensure_data_dirs()
bootstrap_bd_from_source()
load_dotenv(os.path.join(os.path.dirname(__file__), '.env'))
logging.basicConfig(filename=logs_path('error.log'), level=logging.INFO,
                    format='%(asctime)s %(levelname)s %(message)s')
app = Flask(__name__)
app.secret_key = 'AldoAbril1978'
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SESSION_FILE_DIR'] = SESSION_DIR
Session(app)
cache = Cache(app, config={'CACHE_TYPE': 'simple'})

TRANSFER = "TRANSFER"
PROVIDERS = "PROVIDER"
INTERBANK = "INTERBAN"
CUENTA = "MOVIMIENT"
MOVIMIENTOS = "MOVIMIENTOS"
ASIENTO= "EXPORT"


def _general_settings_path():
    return files_path('general_settings.json')


def _account_features_path():
    return files_path('account_features.json')


def _profile_photo_dir():
    return files_path('profile')


def _profile_photo_path(filename='profile_photo.png'):
    return os.path.join(_profile_photo_dir(), filename)


def _profile_photo_context():
    settings = load_general_settings()
    photo_version = int(settings.get('perfil', {}).get('foto_version', 0) or 0)
    return {
        'has_profile_photo': os.path.exists(_profile_photo_path()),
        'profile_photo_url': f"/foto_perfil_actual?v={photo_version}"
    }


def _require_worker_microsoft_login():
    session['system_authenticated'] = True
    session['smtp_authenticated'] = True
    session['quick_password_verified'] = True
    return None


def load_general_settings():
    default_payload = {
        'inicio': {
            'mostrar_resumen': True,
            'notificaciones_activas': True
        },
        'perfil': {
            'foto_version': 0,
            'correo_personal': ''
        },
        'contactos': {
            'correo_soporte': '',
            'correo_copia': ''
        },
        'microsoft365': {
            'tenant_id': 'common',
            'client_id': '',
            'client_secret': '',
            'scope': 'https://graph.microsoft.com/User.Read openid profile email'
        }
    }

    settings_path = _general_settings_path()
    if not os.path.exists(settings_path):
        return default_payload

    try:
        with open(settings_path, 'r', encoding='utf-8') as settings_file:
            payload = json.load(settings_file)

        if not isinstance(payload, dict):
            return default_payload

        merged = default_payload.copy()
        for section_key in default_payload.keys():
            section_value = payload.get(section_key, {})
            if isinstance(section_value, dict):
                merged_section = default_payload[section_key].copy()
                merged_section.update(section_value)
                merged[section_key] = merged_section
        return merged
    except Exception as ex:
        logging.error(f'No se pudo leer configuración general: {ex}')
        return default_payload


def save_general_settings(payload):
    with open(_general_settings_path(), 'w', encoding='utf-8') as settings_file:
        json.dump(payload, settings_file, ensure_ascii=False)


def load_account_features():
    default_payload = {
        'password': {
            'value_encrypted': '',
            'updated_at': ''
        },
        'devices': [
            {
                'id': 1,
                'name': 'Equipo principal',
                'location': 'Oficina',
                'active': True
            }
        ],
        'activity': [],
        'sessions': [],
        'security': {
            'skip_password_when_possible': False,
            'enhanced_browsing': False
        }
    }

    features_path = _account_features_path()
    if not os.path.exists(features_path):
        return default_payload

    try:
        with open(features_path, 'r', encoding='utf-8') as features_file:
            payload = json.load(features_file)

        if not isinstance(payload, dict):
            return default_payload

        merged = default_payload.copy()
        for key in ['password', 'devices', 'activity', 'sessions', 'security']:
            value = payload.get(key)
            if key == 'password' and isinstance(value, dict):
                password_payload = default_payload['password'].copy()
                password_payload.update(value)
                merged['password'] = password_payload
            elif key == 'devices' and isinstance(value, list):
                merged['devices'] = value
            elif key == 'activity' and isinstance(value, list):
                merged['activity'] = value
            elif key == 'sessions' and isinstance(value, list):
                merged['sessions'] = value
            elif key == 'security' and isinstance(value, dict):
                security_payload = default_payload['security'].copy()
                security_payload.update(value)
                merged['security'] = security_payload

        return merged
    except Exception as ex:
        logging.error(f'No se pudo leer account_features: {ex}')
        return default_payload


def save_account_features(payload):
    with open(_account_features_path(), 'w', encoding='utf-8') as features_file:
        json.dump(payload, features_file, ensure_ascii=False)


def add_account_activity(action, detail=''):
    return


def _ensure_current_session_tracked():
    return


@app.before_request
def _before_every_request_account_tracking():
    try:
        _ensure_current_session_tracked()
    except Exception as ex:
        logging.warning(f'No se pudo actualizar sesión de dispositivo: {ex}')

def _enforce_step_flow(current_step: str):
    return None


def _require_access(step=None):
    def _decorator(func):
        @wraps(func)
        def _wrapped(*args, **kwargs):
            auth_redirect = _require_worker_microsoft_login()
            if auth_redirect is not None:
                return auth_redirect

            if step:
                flow_redirect = _enforce_step_flow(step)
                if flow_redirect is not None:
                    return flow_redirect

            return func(*args, **kwargs)

        return _wrapped

    return _decorator

@app.route('/', methods=['POST','GET'])
@_require_access('home')
def home():
    if request.method == 'GET' and request.args.get('reset') == '1':
        session.pop('home_processing_result', None)
        session.pop('home_success_message', None)

    general_settings = load_general_settings()
    config_message = session.pop('config_message', None)
    quick_password_message = session.pop('quick_password_message', None)
    open_mail_settings = bool(session.pop('open_mail_settings', False)) or bool(quick_password_message)
    photo_path = _profile_photo_path()
    has_profile_photo = os.path.exists(photo_path)
    profile_photo_version = general_settings.get('perfil', {}).get('foto_version', 0)

    if request.method == 'POST':
        files = request.files.getlist('file')
        filtered_files = [x for x in files if x.filename!=""]
        if len(filtered_files) <= 1:
            return render_template(
                'home.html',
                error_message='Debe subir por lo menos un archivo.',
                processing_result=session.get('home_processing_result'),
                mensaje_exito=config_message or session.get('home_success_message'),
                quick_password_message=quick_password_message,
                open_mail_settings=open_mail_settings,
                has_profile_photo=has_profile_photo,
                profile_photo_url=f"/foto_perfil_actual?v={profile_photo_version}"
            )
        else:
            try:
                accountService = AccountService()    
                transferService = TransferService()
                interbankService = InterbankService()
                providerService = ProviderService()
                for file in files:
                    nombre =  file.filename.upper() 
                    if (nombre != ""):
                        if CUENTA in nombre:
                            accountService.process_movements(file)
                        elif TRANSFER in nombre:  
                            transferService.setMovimientos(accountService.movimientos)
                            transferService.process_transfers(file)
                        elif INTERBANK in nombre:
                            interbankService.setMovimientos(accountService.movimientos)
                            interbankService.process_interbanks(file)
                        elif PROVIDERS in nombre:
                            providerService.setMovimientos(accountService.movimientos)
                            providerService.process_providers(file)    
                        else:
                            raise Exception("Archivo no ubicado: "+nombre)    
                guardaMovimientos(accountService.movimientos)
                guardaRecaudos(accountService.recaudos)
                resumen = {"movements": accountService.error, "providers": providerService.error, "transfers": transferService.error, "interbanks": interbankService.error}
                session['required_next_step'] = 'asiento'
                session['home_processing_result'] = resumen
                session['home_success_message'] = 'Proceso exitosamente.'
                return redirect(url_for('home'))
    
                #return redirect(url_for('upload'))
            except Exception as e:
                error_message = str(e)
                return render_template(
                    'home.html',
                    error_message=error_message,
                    processing_result=session.get('home_processing_result'),
                    mensaje_exito=config_message or session.get('home_success_message'),
                    quick_password_message=quick_password_message,
                    open_mail_settings=open_mail_settings,
                    has_profile_photo=has_profile_photo,
                    profile_photo_url=f"/foto_perfil_actual?v={profile_photo_version}"
                )
    else:
        return render_template(
            'home.html',
            processing_result=session.get('home_processing_result'),
            mensaje_exito=config_message or session.get('home_success_message'),
            quick_password_message=quick_password_message,
            open_mail_settings=open_mail_settings,
            has_profile_photo=has_profile_photo,
            profile_photo_url=f"/foto_perfil_actual?v={profile_photo_version}"
        )


@app.route('/menu', methods=['GET'])
@_require_access()
def menu():
    tab = request.args.get('tab', 'home').strip().lower()
    tab_map = {
        'home': '/',
        'basedatos': '/basedatos',
        'asiento': '/asiento',
        'correos': '/correos'
    }
    if tab not in tab_map:
        tab = 'home'

    return render_template('menu.html', active_tab=tab, tab_map=tab_map, **_profile_photo_context())

def guardaMovimientos(movimientos):
    movimientos["Fecha"] = pd.to_datetime(movimientos["Fecha"], format="%d/%m/%Y").dt.strftime("%d/%m/%Y")
    excel_file = BytesIO()
    movimientos.to_excel(excel_file, index=False)
    excel_file.seek(0)
    workbook = openpyxl.load_workbook(excel_file)
    worksheet = workbook.active 
    worksheet.column_dimensions["A"].width = 20  
    worksheet.column_dimensions["C"].width = 30  
    worksheet.column_dimensions["K"].width = 40  
    worksheet.column_dimensions["L"].width = 40  
    worksheet.column_dimensions["M"].width = 35 
    worksheet.column_dimensions["N"].width = 40
    for col in worksheet.iter_cols(min_row=1, max_row=1):
        header_value = col[0].value
        if header_value == 'Correo':
            worksheet.column_dimensions[col[0].column_letter].width = 42
            break
    ruta_archivo = files_path('movimientos.xlsx')
    workbook.save(ruta_archivo)

def guardaRecaudos(recaudos):
    excel_file = BytesIO()
    recaudos.to_excel(excel_file, index=False)
    excel_file.seek(0)
    workbook = openpyxl.load_workbook(excel_file)
    worksheet = workbook.active 
    worksheet.column_dimensions["A"].width = 15  
    worksheet.column_dimensions["B"].width = 50  
    worksheet.column_dimensions["C"].width = 25  
    worksheet.column_dimensions["D"].width = 10  
    worksheet.column_dimensions["E"].width = 15 
    worksheet.column_dimensions["F"].width = 10 

    ruta_archivo = files_path('recaudos.xlsx')
    workbook.save(ruta_archivo)

@app.route('/basedatos', methods=['POST','GET'])
@_require_access('basedatos')
def basedatos():
    profile_photo_context = _profile_photo_context()
    config_message = session.pop('config_message', None)
    quick_password_message = session.pop('quick_password_message', None)
    open_mail_settings = bool(session.pop('open_mail_settings', False)) or bool(quick_password_message)

    if request.method == 'POST':
        files    = request.files.getlist('file')
        
        try:
            filtered_files = [x for x in files if x.filename!=""]
                        
            if len(filtered_files) < 1:
                return render_template('base-datos.html', error_message= 'Debe subir por lo menos un archivo.', **profile_photo_context)

            nombres = [f.filename.upper() for f in filtered_files]
            valid_patterns = ['RECAUDO', 'PREPAGO', 'TRABAJADOR', 'CLIENTE']
            invalid_files = [n for n in nombres if not any(pattern in n for pattern in valid_patterns)]
            if invalid_files:
                return render_template('base-datos.html', error_message='Archivo(s) no reconocido(s): ' + ', '.join(invalid_files), **profile_photo_context)

            mensaje_exito = 'Proceso exitosamente.'
            
            base_datos_service = BaseDatosService()  
            base_datos_service.GuardarAchivos(files)  
            session['required_next_step'] = 'home'
            return render_template('base-datos.html', mensaje_exito=mensaje_exito, **profile_photo_context)
                
        except Exception as e:
            error_message = str(e)
            return render_template('base-datos.html', error_message= error_message, **profile_photo_context)

    else:
        return render_template(
            'base-datos.html',
            mensaje_exito=config_message,
            quick_password_message=quick_password_message,
            open_mail_settings=open_mail_settings,
            **profile_photo_context
        )
    
@app.route('/asiento', methods=['POST'])
@_require_access('asiento')
def asiento_procesar():
    profile_photo_context = _profile_photo_context()

    logging.error('asiento_procesar: start')
    files = request.files.getlist('file')
    logging.error('asiento_procesar: received %d files', len(files))
    filtered_files = [x for x in files if x.filename!=""]
    logging.error('asiento_procesar: filtered %d files', len(filtered_files))
    if len(filtered_files) <= 1:
        logging.error('asiento_procesar: not enough files, returning form')
        return render_template('asiento.html', error_message= 'Debe subir ambos archivo.', **profile_photo_context)
    else:
        try:
            asientoService = AsientoService()    
            # detect files: movimientos y asientos
            movimientosfile = None
            asientosfile = None
            for file in files:
                nombre =  file.filename.upper()
                if (nombre != ""):
                    if MOVIMIENTOS in nombre or "MOVIMIENTO" in nombre:
                        movimientosfile = file
                    elif ASIENTO in nombre:
                        asientosfile = file
                    else:
                        # ignore unknown files for now
                        pass

            if movimientosfile is None or asientosfile is None:
                raise Exception('Faltan archivos requeridos: Movimientos o Asientos')

            clientes_email_map = correo_service.load_clientes_email_map_from_bd()

            asientoService.conciliar(movimientosfile, asientosfile)
            #solo si hay asientos se completa en el cache
            if asientoService.df_movimientos is not None:
                emails = correo_service.collect_emails_without_voucher_using_clientes(asientoService.df_movimientos, clientes_email_map)
                guardaAsientos(asientoService.df_movimientos)
                
                # PREPARAR INFORMACIÓN DE VOUCHERS EN HTML (SIN GENERAR PDF)
                try:
                    vouchers_generados = correo_service.build_voucher_data_without_pdf(
                        asientoService.df_movimientos, 
                        clientes_email_map
                    )
                    session['vouchers_generados'] = vouchers_generados
                    logging.info(f'Vouchers HTML preparados: {len(vouchers_generados)}')
                except Exception as ve:
                    logging.error(f'Error preparando vouchers HTML: {str(ve)}')
                    session['vouchers_generados'] = []
                
                # guardar en session para uso posterior y redirigir al flujo de correos
                sorted_emails = sorted(set(emails))
                session['asiento_emails'] = sorted_emails
                session['correos_ready'] = True
                correo_service.save_emails_cache(sorted_emails)
                if len(sorted_emails) == 0:
                    session['asiento_email_warning'] = 'No se encontraron correos para líneas sin voucher. Verifique Referencia y CORREO DE CONTACTO en Base de Datos > Clientes.'
                else:
                    session.pop('asiento_email_warning', None)
                session['required_next_step'] = 'correos'
                return redirect(url_for('asiento_get', resultado_correo=1))
            else: 
                #si hubiera error se pinta la misma pagina y no se redirecciona
                return render_template('asiento.html', error_message= 'No se encontro ningun asiento en el proceso', **profile_photo_context)       
            
        except Exception as e:
            error_message = str(e)
            logging.error('asiento_procesar: exception: %s', error_message)
            return render_template('asiento.html', error_message= error_message, **profile_photo_context)
    # Fallback: ensure the view always returns a response
    logging.error('asiento_procesar: reached end of function without explicit return')
    return render_template('asiento.html', error_message='Error inesperado en el procesamiento', **profile_photo_context)



@app.route('/asiento', methods=['GET'])
@_require_access('asiento')
def asiento_get():
    show_result_mode = request.args.get('resultado_correo', '0') == '1'
    asiento_emails = session.get('asiento_emails', [])
    vouchers_generados = session.get('vouchers_generados', [])
    secure_smtp = correo_service.load_secure_smtp_credentials()
    secure_sender = str(secure_smtp.get('sender', '')).strip()
    mail_sender = str(session.get('worker_sender', '')).strip() or secure_sender
    mail_sender_local = mail_sender.split('@', 1)[0].strip() if mail_sender else ''
    mail_cc = str(session.get('worker_cc', '')).strip() or str(secure_smtp.get('cc', '')).strip()
    page_message = session.pop('config_message', None)
    quick_password_message = session.pop('quick_password_message', None)
    open_mail_settings = bool(session.pop('open_mail_settings', False)) or bool(quick_password_message)

    return render_template(
        'asiento.html',
        show_result_mode=show_result_mode,
        asiento_emails=asiento_emails,
        total_vouchers=len(vouchers_generados),
        mail_sender=mail_sender_local,
        mail_cc=mail_cc,
        mensaje_exito=page_message,
        quick_password_message=quick_password_message,
        open_mail_settings=open_mail_settings,
        **_profile_photo_context()
    )


correo_service.register_correo_routes(
    app=app,
    require_worker_microsoft_login_func=_require_worker_microsoft_login,
    enforce_step_flow_func=_enforce_step_flow,
    load_general_settings_func=load_general_settings,
    save_general_settings_func=save_general_settings,
    profile_photo_path_func=_profile_photo_path,
    add_account_activity_func=add_account_activity,
)


@app.route('/foto_perfil_actual', methods=['GET'])
def foto_perfil_actual():
    photo_path = _profile_photo_path()
    if not os.path.exists(photo_path):
        return '', 404
    return send_file(photo_path)


@app.route('/favicon.ico', methods=['GET'])
def favicon():
    favicon_path = os.path.join(app.static_folder, 'images', 'favicon', 'favicon.ico')
    if not os.path.exists(favicon_path):
        return '', 404
    return send_file(favicon_path)

@app.route('/upload', methods=['POST','GET'])
def upload():
    
    if request.method == 'POST':
        ruta_archivo = files_path('movimientos.xlsx')
        return send_file(ruta_archivo, as_attachment=True, download_name="movimientos.xlsx")
   
    

@app.route('/download_recaudos', methods=['POST'])
def download_recaudos():
    ruta_archivo = files_path('recaudos.xlsx')
    return send_file(ruta_archivo, as_attachment=True, download_name="recaudos.xlsx")

def guardaAsientos(movimientosAsientos):
    movimientosAsientos["Fecha"] = pd.to_datetime(movimientosAsientos["Fecha"], format="%d/%m/%Y").dt.strftime("%d/%m/%Y")
    excel_file = BytesIO()
    movimientosAsientos.to_excel(excel_file, index=False)
    excel_file.seek(0)
    workbook = openpyxl.load_workbook(excel_file)
    worksheet = workbook.active 
    worksheet.column_dimensions["A"].width = 20  
    worksheet.column_dimensions["C"].width = 30  
    worksheet.column_dimensions["K"].width = 40  
    worksheet.column_dimensions["L"].width = 40  
    worksheet.column_dimensions["M"].width = 35 
    worksheet.column_dimensions["N"].width = 28 

    ruta_archivo = files_path('asientos.xlsx')
    workbook.save(ruta_archivo)


@app.route('/download_asientos', methods=['GET', 'POST'])
def dowload_asientos():
    ruta_archivo = files_path('asientos.xlsx')
    return send_file(ruta_archivo, as_attachment=True, download_name="Asiento.xlsx")


@app.route('/cerrar_sesion', methods=['GET'])
def cerrar_sesion():
    session.clear()
    return redirect(url_for('iniciar_sesion'))



@app.errorhandler(Exception)
def handle_exception(e):
     # Log full traceback to file
     tb = traceback.format_exc()
     logging.error('Unhandled exception:\n%s', tb)
     # return a friendly error page
     return render_template('error.html', message=str(e)), 500


if __name__ == '__main__':
    # Configuración para red local
    app.run(host='0.0.0.0', port=5000, debug=True)

    