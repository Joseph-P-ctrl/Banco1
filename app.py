from flask import Flask, jsonify, render_template, request, send_file, session, redirect, url_for
from AccountService import AccountService
from InterbankService import InterbankService
from ProviderService import ProviderService
from TransferService import TransferService
from BaseDatosService import BaseDatosService
from AsientoService import AsientoService
from VoucherService import VoucherService
from io import BytesIO
from flask_session import Session
from flask_caching import Cache
from openpyxl import load_workbook
import pandas as pd
import os
import openpyxl
from openpyxl import Workbook
import re
import logging
import traceback
import json
import smtplib
import uuid
import urllib.parse
import urllib.request
from urllib.error import HTTPError, URLError
from email.message import EmailMessage
from dotenv import load_dotenv
from cryptography.fernet import Fernet
from storage_paths import ensure_data_dirs, bootstrap_bd_from_source, files_path, logs_path, SESSION_DIR, bd_path, vouchers_path

# setup logging
ensure_data_dirs()
bootstrap_bd_from_source()
load_dotenv(os.path.join(os.path.dirname(__file__), '.env'))
logging.basicConfig(filename=logs_path('error.log'), level=logging.INFO,
                    format='%(asctime)s %(levelname)s %(message)s')
app = Flask(__name__)
app.secret_key = 'AldoAbril1978'
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SESSION_FILE_DIR'] = SESSION_DIR
Session(app)
cache = Cache(app, config={'CACHE_TYPE': 'simple'})

TRANSFER = "TRANSFER"
PROVIDERS = "PROVIDER"
INTERBANK = "INTERBAN"
CUENTA = "MOVIMIENT"
MOVIMIENTOS = "MOVIMIENTOS"
ASIENTO= "EXPORT"

COMPANY_KEYWORDS = {
    'SAC', 'S.A.C', 'SRL', 'S.R.L', 'SA', 'S.A', 'EIRL', 'E.I.R.L',
    'GERENCIA', 'DIRECCION', 'DIREC', 'REGIONAL', 'MUNICIPALIDAD',
    'MINISTERIO', 'GOBIERNO', 'UNIDAD', 'LOGISTICA', 'AGRICULTURA',
    'POLICIAL', 'HOSPITAL', 'UNIVERSIDAD', 'COLEGIO', 'EMPRESA',
    'SERVICIOS', 'AREA', 'OFICINA'
}


def _smtp_key_path():
    return files_path('smtp_credentials.key')


def _smtp_credentials_path():
    return files_path('smtp_credentials.json')


def _google_config_path():
    return files_path('google_config.json')


def _general_settings_path():
    return files_path('general_settings.json')


def _account_features_path():
    return files_path('account_features.json')


def _profile_photo_dir():
    return files_path('profile')


def _profile_photo_path(filename='profile_photo.png'):
    return os.path.join(_profile_photo_dir(), filename)


def _allowed_image_extension(filename):
    lower_name = str(filename or '').strip().lower()
    return lower_name.endswith('.png') or lower_name.endswith('.jpg') or lower_name.endswith('.jpeg') or lower_name.endswith('.webp')


def _safe_user_agent():
    try:
        return str(request.user_agent.string or '').strip()
    except Exception:
        return ''


def _client_ip():
    forwarded_for = request.headers.get('X-Forwarded-For', '').strip()
    if forwarded_for:
        return forwarded_for.split(',')[0].strip()
    return str(request.remote_addr or '').strip() or 'desconocido'


def _detect_os_family(user_agent_text):
    ua = str(user_agent_text or '').lower()
    if 'android' in ua:
        return 'Android'
    if 'windows' in ua:
        return 'Windows'
    if 'iphone' in ua or 'ipad' in ua or 'ios' in ua:
        return 'iOS'
    if 'mac os' in ua or 'macintosh' in ua:
        return 'macOS'
    if 'linux' in ua:
        return 'Linux'
    return 'Otro'


def _detect_device_type(user_agent_text):
    ua = str(user_agent_text or '').lower()
    if 'android' in ua or 'iphone' in ua or 'mobile' in ua:
        return 'Móvil'
    if 'ipad' in ua or 'tablet' in ua:
        return 'Tablet'
    return 'Computadora'


def _relative_time_label(timestamp_value):
    if not timestamp_value:
        return 'Sin registro'
    try:
        last_seen = pd.to_datetime(timestamp_value, format='%d/%m/%Y %H:%M:%S', errors='coerce')
        if pd.isna(last_seen):
            return str(timestamp_value)
        now_value = pd.Timestamp.now()
        diff = now_value - last_seen
        total_seconds = max(int(diff.total_seconds()), 0)

        if total_seconds < 60:
            return 'Hace unos segundos'
        if total_seconds < 3600:
            minutes = total_seconds // 60
            return f'Hace {minutes} min'
        if total_seconds < 86400:
            hours = total_seconds // 3600
            return f'Hace {hours} hora(s)'
        days = total_seconds // 86400
        return f'Hace {days} día(s)'
    except Exception:
        return str(timestamp_value)


def _get_fernet():
    env_key = os.environ.get('OUTLOOK_CREDENTIALS_KEY', '').strip()
    if env_key:
        key = env_key.encode('utf-8')
        return Fernet(key)

    key_path = _smtp_key_path()
    if os.path.exists(key_path):
        with open(key_path, 'rb') as key_file:
            key = key_file.read().strip()
    else:
        key = Fernet.generate_key()
        with open(key_path, 'wb') as key_file:
            key_file.write(key)

    return Fernet(key)


def normalize_sender_email(sender_value):
    sender_clean = str(sender_value or '').strip()
    sender_lower = sender_clean.lower()
    typo_domain = '@distrluz.com.pe'
    corrected_domain = '@distriluz.com.pe'

    if sender_lower.endswith(typo_domain):
        corrected = sender_clean[:-len(typo_domain)] + corrected_domain
        return corrected, True

    return sender_clean, False


def _build_display_name_from_email(email_value):
    email_text = str(email_value or '').strip().lower()
    if not email_text or '@' not in email_text:
        return 'Usuario'
    local_part = email_text.split('@', 1)[0]
    tokens = [token for token in re.split(r'[._\-]+', local_part) if token and not token.isdigit()]
    if not tokens:
        return 'Usuario'
    return ' '.join(token.capitalize() for token in tokens[:4])


def _normalize_smtp_security(security_value):
    normalized = str(security_value or '').strip().lower()
    if normalized not in ('ssl', 'starttls', 'auto'):
        return 'starttls'
    return normalized


def _parse_smtp_port(port_value):
    try:
        return int(str(port_value or '').strip())
    except Exception:
        return 587


def _should_retry_with_port_25(smtp_port, conn_ex):
    if int(smtp_port or 0) != 587:
        return False

    error_text = str(conn_ex or '').lower()
    return any(token in error_text for token in (
        'winerror 10061',
        'connection refused',
        'actively refused',
        'timed out',
        'timeout',
    ))


def _require_worker_microsoft_login():
    session['system_authenticated'] = True
    session['smtp_authenticated'] = True
    session['quick_password_verified'] = True
    return None


MAX_LOGIN_ATTEMPTS = 5
LOGIN_BLOCK_MINUTES = 10


def _get_login_lock_state():
    block_until_raw = str(session.get('login_block_until', '')).strip()
    if not block_until_raw:
        return False, 0

    block_until = pd.to_datetime(block_until_raw, errors='coerce')
    if pd.isna(block_until):
        session.pop('login_block_until', None)
        return False, 0

    now_value = pd.Timestamp.now()
    if now_value >= block_until:
        session.pop('login_block_until', None)
        session['login_attempts'] = 0
        return False, 0

    remaining_minutes = int(max((block_until - now_value).total_seconds(), 0) // 60) + 1
    return True, remaining_minutes


def _register_login_failure(base_message):
    attempts = int(session.get('login_attempts', 0) or 0) + 1
    session['login_attempts'] = attempts

    remaining = max(MAX_LOGIN_ATTEMPTS - attempts, 0)
    if remaining <= 0:
        block_until = pd.Timestamp.now() + pd.Timedelta(minutes=LOGIN_BLOCK_MINUTES)
        session['login_block_until'] = block_until.strftime('%Y-%m-%d %H:%M:%S')
        session['login_attempts'] = 0
        session['login_message'] = (
            f'{base_message} Alcanzaste el máximo de {MAX_LOGIN_ATTEMPTS} intentos. '
            f'Intenta nuevamente en {LOGIN_BLOCK_MINUTES} minutos.'
        )
        return

    session['login_message'] = f'{base_message} Intentos restantes: {remaining}.'


def _reset_login_failures():
    session['login_attempts'] = 0
    session.pop('login_block_until', None)


def _validate_microsoft365_api_login(username, password):
    settings = load_general_settings()
    microsoft365_settings = settings.get('microsoft365', {}) if isinstance(settings, dict) else {}

    tenant_id = (
        os.environ.get('M365_TENANT_ID', '').strip()
        or os.environ.get('MICROSOFT_TENANT_ID', '').strip()
        or str(microsoft365_settings.get('tenant_id', '')).strip()
        or 'common'
    )
    client_id = (
        os.environ.get('M365_CLIENT_ID', '').strip()
        or os.environ.get('MICROSOFT_CLIENT_ID', '').strip()
        or str(microsoft365_settings.get('client_id', '')).strip()
    )
    client_secret = (
        os.environ.get('M365_CLIENT_SECRET', '').strip()
        or os.environ.get('MICROSOFT_CLIENT_SECRET', '').strip()
        or str(microsoft365_settings.get('client_secret', '')).strip()
    )
    scope = (
        os.environ.get('M365_SCOPE', '').strip()
        or os.environ.get('MICROSOFT_SCOPE', '').strip()
        or str(microsoft365_settings.get('scope', '')).strip()
        or 'https://graph.microsoft.com/User.Read openid profile email'
    )

    if not client_id:
        return False, 'API Microsoft 365 no configurada. Falta M365_CLIENT_ID (o MICROSOFT_CLIENT_ID).'

    token_url = f'https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token'
    payload = {
        'client_id': client_id,
        'grant_type': 'password',
        'username': username,
        'password': password,
        'scope': scope
    }
    if client_secret:
        payload['client_secret'] = client_secret

    encoded_payload = urllib.parse.urlencode(payload).encode('utf-8')
    token_request = urllib.request.Request(
        token_url,
        data=encoded_payload,
        headers={'Content-Type': 'application/x-www-form-urlencoded'}
    )

    try:
        with urllib.request.urlopen(token_request, timeout=30) as token_response:
            token_data = json.loads(token_response.read().decode('utf-8'))
    except HTTPError as http_ex:
        try:
            error_payload = json.loads(http_ex.read().decode('utf-8'))
            error_text = error_payload.get('error_description') or error_payload.get('error') or str(http_ex)
        except Exception:
            error_text = str(http_ex)
        return False, f'Error API Microsoft 365: {error_text}'
    except URLError as url_ex:
        return False, f'No se pudo conectar a Microsoft 365 API: {url_ex}'
    except Exception as ex:
        return False, f'Error validando API Microsoft 365: {ex}'

    access_token = str(token_data.get('access_token', '')).strip()
    if not access_token:
        return False, 'Microsoft 365 API no devolvió token de acceso para esta cuenta.'

    me_request = urllib.request.Request(
        'https://graph.microsoft.com/v1.0/me',
        headers={'Authorization': f'Bearer {access_token}'}
    )

    try:
        with urllib.request.urlopen(me_request, timeout=20) as me_response:
            me_data = json.loads(me_response.read().decode('utf-8'))
            if not me_data.get('id'):
                return False, 'Microsoft 365 API respondió sin identidad de usuario.'
    except Exception as ex:
        return False, f'No se pudo validar perfil en Microsoft 365 API: {ex}'

    return True, ''


def _is_microsoft365_graph_configured():
    settings = load_general_settings()
    microsoft365_settings = settings.get('microsoft365', {}) if isinstance(settings, dict) else {}

    client_id = (
        os.environ.get('M365_CLIENT_ID', '').strip()
        or os.environ.get('MICROSOFT_CLIENT_ID', '').strip()
        or str(microsoft365_settings.get('client_id', '')).strip()
    )
    return bool(client_id)


def _validate_smtp_login(sender, password, smtp_host, smtp_port, smtp_security):
    used_security = _normalize_smtp_security(smtp_security)
    smtp_conn = None

    try:
        if used_security == 'ssl':
            smtp_conn = smtplib.SMTP_SSL(smtp_host, smtp_port, timeout=30)
            smtp_conn.ehlo()
        else:
            smtp_conn = smtplib.SMTP(smtp_host, smtp_port, timeout=30)
            smtp_conn.ehlo()
            smtp_conn.starttls()
            smtp_conn.ehlo()
    except Exception as conn_ex:
        wrong_version = 'WRONG_VERSION_NUMBER' in str(conn_ex).upper()
        can_retry_starttls = used_security in ('ssl', 'auto') and wrong_version
        if can_retry_starttls:
            smtp_conn = smtplib.SMTP(smtp_host, smtp_port, timeout=30)
            smtp_conn.ehlo()
            smtp_conn.starttls()
            smtp_conn.ehlo()
            used_security = 'starttls'
        elif _should_retry_with_port_25(smtp_port, conn_ex):
            smtp_conn = smtplib.SMTP(smtp_host, 25, timeout=30)
            smtp_conn.ehlo()
            smtp_conn.starttls()
            smtp_conn.ehlo()
            used_security = 'starttls'
        else:
            return False, sender, used_security, f'No se pudo conectar al servidor SMTP ({smtp_host}:{smtp_port}): {conn_ex}'

    with smtp_conn as smtp:
        try:
            smtp.login(sender, password)
            return True, sender, used_security, ''
        except smtplib.SMTPAuthenticationError:
            corrected_sender, sender_corrected = normalize_sender_email(sender)
            if sender_corrected and corrected_sender != sender:
                try:
                    smtp.login(corrected_sender, password)
                    return True, corrected_sender, used_security, ''
                except smtplib.SMTPAuthenticationError:
                    pass

            return (
                False,
                sender,
                used_security,
                'No se pudieron validar tus credenciales de correo. Verifica usuario y contraseña e inténtalo nuevamente.'
            )
        except Exception as login_ex:
            return False, sender, used_security, f'No se pudo autenticar en SMTP: {login_ex}'


def _validate_office365_and_smtp_login(sender, password, smtp_host, smtp_port, smtp_security):
    smtp_ok, validated_sender, used_security, smtp_error = _validate_smtp_login(
        sender,
        password,
        smtp_host,
        smtp_port,
        smtp_security
    )
    if not smtp_ok:
        return False, validated_sender, used_security, smtp_error

    graph_configured = _is_microsoft365_graph_configured()
    if not graph_configured:
        session['worker_login_via_graph'] = False
        session['worker_graph_verified'] = False
        session['worker_auth_method'] = 'SMTP OWA'
        session['worker_graph_warning'] = 'Microsoft 365 Graph no está configurado. Se inició sesión con SMTP.'
        return True, validated_sender, used_security, ''

    graph_ok, graph_error = _validate_microsoft365_api_login(validated_sender, password)
    if not graph_ok:
        session['worker_login_via_graph'] = False
        session['worker_graph_verified'] = False
        session['worker_auth_method'] = 'SMTP OWA'
        session['worker_graph_warning'] = 'Conexión SMTP correcta, pero Microsoft 365 no validó: ' + graph_error
        return True, validated_sender, used_security, ''

    session['worker_login_via_graph'] = True
    session['worker_graph_verified'] = True
    session.pop('worker_graph_warning', None)
    session['worker_auth_method'] = 'SMTP OWA + Microsoft 365 Graph'
    return True, validated_sender, used_security, ''


def save_secure_smtp_credentials(sender_value, password_value, smtp_host_value=None, smtp_port_value=None, smtp_security_value=None):
    fernet = _get_fernet()
    payload = {
        'sender_encrypted': fernet.encrypt(sender_value.encode('utf-8')).decode('utf-8'),
        'password_encrypted': fernet.encrypt(password_value.encode('utf-8')).decode('utf-8')
    }
    if smtp_host_value is not None:
        payload['smtp_host'] = str(smtp_host_value).strip()
    if smtp_port_value is not None:
        payload['smtp_port'] = str(smtp_port_value).strip()
    if smtp_security_value is not None:
        payload['smtp_security'] = str(smtp_security_value).strip().lower()
    with open(_smtp_credentials_path(), 'w', encoding='utf-8') as out_file:
        json.dump(payload, out_file, ensure_ascii=False)


def load_secure_smtp_credentials():
    cred_path = _smtp_credentials_path()
    if not os.path.exists(cred_path):
        return {}

    try:
        with open(cred_path, 'r', encoding='utf-8') as in_file:
            payload = json.load(in_file)

        sender_encrypted = payload.get('sender_encrypted', '')
        password_encrypted = payload.get('password_encrypted', '')
        if not sender_encrypted or not password_encrypted:
            return {}

        fernet = _get_fernet()
        sender_value = fernet.decrypt(sender_encrypted.encode('utf-8')).decode('utf-8')
        password_value = fernet.decrypt(password_encrypted.encode('utf-8')).decode('utf-8')
        return {
            'sender': sender_value,
            'password': password_value,
            'smtp_host': str(payload.get('smtp_host', '')).strip(),
            'smtp_port': str(payload.get('smtp_port', '')).strip(),
            'smtp_security': str(payload.get('smtp_security', '')).strip().lower()
        }
    except Exception as ex:
        logging.error(f'No se pudo leer credenciales SMTP cifradas: {ex}')
        return {}


def save_google_config(config_values):
    payload = {
        'project_id': str(config_values.get('project_id', '')).strip(),
        'client_id': str(config_values.get('client_id', '')).strip(),
        'redirect_uri': str(config_values.get('redirect_uri', '')).strip(),
        'scopes': str(config_values.get('scopes', '')).strip(),
        'enabled': bool(config_values.get('enabled', False))
    }

    fernet = _get_fernet()

    client_secret = str(config_values.get('client_secret', '')).strip()
    api_key = str(config_values.get('api_key', '')).strip()

    if client_secret:
        payload['client_secret_encrypted'] = fernet.encrypt(client_secret.encode('utf-8')).decode('utf-8')
    if api_key:
        payload['api_key_encrypted'] = fernet.encrypt(api_key.encode('utf-8')).decode('utf-8')

    with open(_google_config_path(), 'w', encoding='utf-8') as out_file:
        json.dump(payload, out_file, ensure_ascii=False)


def load_google_config():
    config_path = _google_config_path()
    if not os.path.exists(config_path):
        return {
            'project_id': '',
            'client_id': '',
            'client_secret': '',
            'redirect_uri': '',
            'scopes': 'openid email profile',
            'api_key': '',
            'enabled': False
        }

    try:
        with open(config_path, 'r', encoding='utf-8') as in_file:
            payload = json.load(in_file)

        fernet = _get_fernet()

        client_secret = ''
        api_key = ''

        client_secret_encrypted = payload.get('client_secret_encrypted', '')
        if client_secret_encrypted:
            client_secret = fernet.decrypt(client_secret_encrypted.encode('utf-8')).decode('utf-8')

        api_key_encrypted = payload.get('api_key_encrypted', '')
        if api_key_encrypted:
            api_key = fernet.decrypt(api_key_encrypted.encode('utf-8')).decode('utf-8')

        return {
            'project_id': str(payload.get('project_id', '')).strip(),
            'client_id': str(payload.get('client_id', '')).strip(),
            'client_secret': client_secret,
            'redirect_uri': str(payload.get('redirect_uri', '')).strip(),
            'scopes': str(payload.get('scopes', 'openid email profile')).strip(),
            'api_key': api_key,
            'enabled': bool(payload.get('enabled', False))
        }
    except Exception as ex:
        logging.error(f'No se pudo leer configuración Google: {ex}')
        return {
            'project_id': '',
            'client_id': '',
            'client_secret': '',
            'redirect_uri': '',
            'scopes': 'openid email profile',
            'api_key': '',
            'enabled': False
        }


def load_general_settings():
    default_payload = {
        'inicio': {
            'mostrar_resumen': True,
            'notificaciones_activas': True
        },
        'perfil': {
            'nombre_mostrar': '',
            'cargo': '',
            'telefono': '',
            'foto_version': 0,
            'genero': '',
            'correo_personal': '',
            'correo_alterno': '',
            'fecha_nacimiento': '',
            'idioma': 'Español (España)',
            'direccion_casa': '',
            'direccion_trabajo': '',
            'otras_direcciones': ''
        },
        'seguridad': {
            'doble_factor': False,
            'cerrar_sesiones': True
        },
        'privacidad': {
            'compartir_datos': False,
            'analytics': False
        },
        'contactos': {
            'correo_soporte': '',
            'correo_copia': ''
        },
        'microsoft365': {
            'tenant_id': 'common',
            'client_id': '',
            'client_secret': '',
            'scope': 'https://graph.microsoft.com/User.Read openid profile email'
        }
    }

    settings_path = _general_settings_path()
    if not os.path.exists(settings_path):
        return default_payload

    try:
        with open(settings_path, 'r', encoding='utf-8') as settings_file:
            payload = json.load(settings_file)

        if not isinstance(payload, dict):
            return default_payload

        merged = default_payload.copy()
        for section_key in default_payload.keys():
            section_value = payload.get(section_key, {})
            if isinstance(section_value, dict):
                merged_section = default_payload[section_key].copy()
                merged_section.update(section_value)
                merged[section_key] = merged_section
        return merged
    except Exception as ex:
        logging.error(f'No se pudo leer configuración general: {ex}')
        return default_payload


def save_general_settings(payload):
    with open(_general_settings_path(), 'w', encoding='utf-8') as settings_file:
        json.dump(payload, settings_file, ensure_ascii=False)


def sync_email_config_to_modules(sender_email):
    sender_clean = str(sender_email or '').strip().lower()
    if not sender_clean:
        return

    settings = load_general_settings()

    if 'perfil' not in settings or not isinstance(settings['perfil'], dict):
        settings['perfil'] = {}
    if 'contactos' not in settings or not isinstance(settings['contactos'], dict):
        settings['contactos'] = {}

    settings['perfil']['correo_personal'] = sender_clean

    soporte_actual = str(settings['contactos'].get('correo_soporte', '')).strip()
    if not soporte_actual:
        settings['contactos']['correo_soporte'] = sender_clean

    save_general_settings(settings)


def get_connected_account_email():
    secure_smtp = load_secure_smtp_credentials()
    sender = str(secure_smtp.get('sender', '')).strip()
    if sender:
        return sender

    settings = load_general_settings()
    perfil = settings.get('perfil', {}) if isinstance(settings, dict) else {}
    return str(perfil.get('correo_personal', '')).strip()


def load_account_features():
    default_payload = {
        'password': {
            'value_encrypted': '',
            'updated_at': ''
        },
        'devices': [
            {
                'id': 1,
                'name': 'Equipo principal',
                'location': 'Oficina',
                'active': True
            }
        ],
        'activity': [],
        'sessions': [],
        'security': {
            'skip_password_when_possible': False,
            'enhanced_browsing': False,
            'recovery_phone': '',
            'recovery_email': '',
            'two_factor_phone': '',
            'backup_codes': 0,
            'device_prompt_count': 0
        },
        'third_party_apps': [],
        'contacts_share': {
            'family': {
                'enabled': False,
                'invites_remaining': 5,
                'group_name': 'Mi grupo familiar'
            },
            'preferences': {
                'sync_interactions': True,
                'sync_device_contacts': True,
                'share_location': False,
                'profile_visible': True
            },
            'contacts': [
                {
                    'id': 1,
                    'name': 'Contacto soporte',
                    'email': '',
                    'phone': '',
                    'source': 'Manual',
                    'blocked': False
                }
            ],
            'profiles': [
                {
                    'id': 1,
                    'name': 'Perfil principal',
                    'description': 'Visible para servicios internos'
                }
            ]
        },
        'privacy_center': {
            'search_personalization': True,
            'play_personalization': True,
            'web_activity': True,
            'play_history': True,
            'youtube_history': True,
            'maps_timeline': False,
            'ads_personalized': True,
            'partner_ads': False,
            'legacy_plan_created': False,
            'delete_account_requested_at': '',
            'service_emails_enabled': True,
            'fit_data_enabled': True,
            'voice_match_enabled': False,
            'download_requests': 0
        }
    }

    features_path = _account_features_path()
    if not os.path.exists(features_path):
        return default_payload

    try:
        with open(features_path, 'r', encoding='utf-8') as features_file:
            payload = json.load(features_file)

        if not isinstance(payload, dict):
            return default_payload

        merged = default_payload.copy()
        for key in ['password', 'devices', 'activity', 'sessions', 'security', 'third_party_apps', 'contacts_share', 'privacy_center']:
            value = payload.get(key)
            if key == 'password' and isinstance(value, dict):
                password_payload = default_payload['password'].copy()
                password_payload.update(value)
                merged['password'] = password_payload
            elif key == 'devices' and isinstance(value, list):
                merged['devices'] = value
            elif key == 'activity' and isinstance(value, list):
                merged['activity'] = value
            elif key == 'sessions' and isinstance(value, list):
                merged['sessions'] = value
            elif key == 'security' and isinstance(value, dict):
                security_payload = default_payload['security'].copy()
                security_payload.update(value)
                merged['security'] = security_payload
            elif key == 'third_party_apps' and isinstance(value, list):
                merged['third_party_apps'] = value
            elif key == 'contacts_share' and isinstance(value, dict):
                contacts_default = default_payload['contacts_share']
                merged_contacts = {
                    'family': contacts_default['family'].copy(),
                    'preferences': contacts_default['preferences'].copy(),
                    'contacts': contacts_default['contacts'][:],
                    'profiles': contacts_default['profiles'][:]
                }

                family_value = value.get('family', {})
                if isinstance(family_value, dict):
                    merged_contacts['family'].update(family_value)

                preferences_value = value.get('preferences', {})
                if isinstance(preferences_value, dict):
                    merged_contacts['preferences'].update(preferences_value)

                contacts_value = value.get('contacts', [])
                if isinstance(contacts_value, list):
                    merged_contacts['contacts'] = contacts_value

                profiles_value = value.get('profiles', [])
                if isinstance(profiles_value, list):
                    merged_contacts['profiles'] = profiles_value

                merged['contacts_share'] = merged_contacts
            elif key == 'privacy_center' and isinstance(value, dict):
                privacy_payload = default_payload['privacy_center'].copy()
                privacy_payload.update(value)
                merged['privacy_center'] = privacy_payload

        return merged
    except Exception as ex:
        logging.error(f'No se pudo leer account_features: {ex}')
        return default_payload


def save_account_features(payload):
    with open(_account_features_path(), 'w', encoding='utf-8') as features_file:
        json.dump(payload, features_file, ensure_ascii=False)


def add_account_activity(action, detail=''):
    payload = load_account_features()
    activity_list = payload.get('activity', [])
    timestamp = pd.Timestamp.now().strftime('%d/%m/%Y %H:%M:%S')
    activity_list.insert(0, {
        'timestamp': timestamp,
        'action': str(action or '').strip(),
        'detail': str(detail or '').strip()
    })

    if len(activity_list) > 200:
        activity_list = activity_list[:200]

    payload['activity'] = activity_list
    save_account_features(payload)


def clear_account_activity(mode='all', keep_last=0):
    payload = load_account_features()
    if mode == 'last' and keep_last > 0:
        payload['activity'] = payload.get('activity', [])[:keep_last]
    else:
        payload['activity'] = []
    save_account_features(payload)


def _ensure_current_session_tracked():
    route_path = request.path or ''
    if route_path.startswith('/static') or route_path.startswith('/foto_perfil_actual'):
        return

    account_session_id = session.get('account_session_id', '')
    if not account_session_id:
        account_session_id = str(uuid.uuid4())
        session['account_session_id'] = account_session_id

    user_agent_text = _safe_user_agent()
    ip_value = _client_ip()
    now_text = pd.Timestamp.now().strftime('%d/%m/%Y %H:%M:%S')

    payload = load_account_features()
    sessions_list = payload.get('sessions', [])

    existing_index = -1
    for index, item in enumerate(sessions_list):
        if str(item.get('session_id', '')).strip() == account_session_id:
            existing_index = index
            break

    if existing_index >= 0:
        current = sessions_list[existing_index]
        current['last_seen'] = now_text
        current['ip'] = ip_value
        current['user_agent'] = user_agent_text
        current['os_family'] = _detect_os_family(user_agent_text)
        current['device_type'] = _detect_device_type(user_agent_text)
        current['active'] = True
        sessions_list[existing_index] = current
    else:
        sessions_list.insert(0, {
            'session_id': account_session_id,
            'first_seen': now_text,
            'last_seen': now_text,
            'ip': ip_value,
            'user_agent': user_agent_text,
            'os_family': _detect_os_family(user_agent_text),
            'device_type': _detect_device_type(user_agent_text),
            'active': True
        })

    sanitized_sessions = []
    now_value = pd.Timestamp.now()
    for item in sessions_list:
        last_seen_text = str(item.get('last_seen', '')).strip()
        is_active = False
        if last_seen_text:
            try:
                parsed_last_seen = pd.to_datetime(last_seen_text, format='%d/%m/%Y %H:%M:%S', errors='coerce')
                if pd.isna(parsed_last_seen):
                    raise ValueError('Fecha inválida')
                delta = now_value - parsed_last_seen
                is_active = delta.total_seconds() <= (7 * 24 * 3600)
            except Exception:
                is_active = bool(item.get('active', False))
        item['active'] = is_active
        sanitized_sessions.append(item)

    payload['sessions'] = sanitized_sessions[:200]
    save_account_features(payload)


@app.before_request
def _before_every_request_account_tracking():
    try:
        _ensure_current_session_tracked()
    except Exception as ex:
        logging.warning(f'No se pudo actualizar sesión de dispositivo: {ex}')


def save_account_password(new_password):
    if not new_password:
        return

    payload = load_account_features()
    fernet = _get_fernet()
    payload['password']['value_encrypted'] = fernet.encrypt(new_password.encode('utf-8')).decode('utf-8')
    payload['password']['updated_at'] = pd.Timestamp.now().strftime('%d/%m/%Y %H:%M:%S')
    save_account_features(payload)


def get_account_password_updated_at():
    payload = load_account_features()
    return str(payload.get('password', {}).get('updated_at', '')).strip()


def get_account_password_value():
    payload = load_account_features()
    encrypted_value = str(payload.get('password', {}).get('value_encrypted', '')).strip()
    if not encrypted_value:
        return ''

    try:
        fernet = _get_fernet()
        return fernet.decrypt(encrypted_value.encode('utf-8')).decode('utf-8')
    except Exception:
        return ''


def _security_activity_items(limit=5):
    payload = load_account_features()
    security_keywords = ['seguridad', 'contraseña', 'sesión', 'correo', 'dispositivo', 'autenticación']

    filtered = []
    for item in payload.get('activity', []):
        action_value = str(item.get('action', '')).strip()
        detail_value = str(item.get('detail', '')).strip()
        haystack = f"{action_value} {detail_value}".lower()
        if any(keyword in haystack for keyword in security_keywords):
            filtered.append(item)
        if len(filtered) >= max(int(limit or 0), 1):
            break
    return filtered


def _group_sessions_summary(sessions_list):
    summary = {}
    for item in sessions_list:
        os_family = str(item.get('os_family', '')).strip() or 'Otro'
        summary[os_family] = summary.get(os_family, 0) + 1
    return summary


def _default_third_party_apps():
    now_text = pd.Timestamp.now().strftime('%d/%m/%Y %H:%M:%S')
    return [
        {'id': 1, 'name': 'Alibaba.com', 'description': 'B2B marketplace', 'connected_at': now_text},
        {'id': 2, 'name': 'AliExpress', 'description': 'Shopping app', 'connected_at': now_text},
        {'id': 3, 'name': 'Canva', 'description': 'Diseño colaborativo', 'connected_at': now_text}
    ]


def _sanitize_int(value, default_value=0):
    try:
        return int(str(value).strip())
    except Exception:
        return int(default_value)


def _deactivate_other_sessions(payload, current_session_id):
    sessions_list = payload.get('sessions', [])
    affected = 0
    current_clean = str(current_session_id or '').strip()

    for item in sessions_list:
        session_id_value = str(item.get('session_id', '')).strip()
        if session_id_value and session_id_value != current_clean and bool(item.get('active', False)):
            item['active'] = False
            affected += 1

    payload['sessions'] = sessions_list
    return affected


def _next_numeric_id(items):
    if not isinstance(items, list):
        return 1
    return max([_sanitize_int(item.get('id', 0), 0) for item in items], default=0) + 1


def _ensure_contacts_share(payload, settings=None):
    if settings is None:
        settings = load_general_settings()

    contacts_share = payload.get('contacts_share', {})
    if not isinstance(contacts_share, dict):
        contacts_share = {}

    family = contacts_share.get('family', {})
    if not isinstance(family, dict):
        family = {}

    preferences = contacts_share.get('preferences', {})
    if not isinstance(preferences, dict):
        preferences = {}

    contacts = contacts_share.get('contacts', [])
    if not isinstance(contacts, list):
        contacts = []

    profiles = contacts_share.get('profiles', [])
    if not isinstance(profiles, list):
        profiles = []

    family_defaults = {
        'enabled': False,
        'invites_remaining': 5,
        'group_name': 'Mi grupo familiar'
    }
    for key, default_value in family_defaults.items():
        if key not in family:
            family[key] = default_value

    pref_defaults = {
        'sync_interactions': True,
        'sync_device_contacts': True,
        'share_location': False,
        'profile_visible': True
    }
    for key, default_value in pref_defaults.items():
        if key not in preferences:
            preferences[key] = default_value

    support_email = str(settings.get('contactos', {}).get('correo_soporte', '')).strip().lower()
    if not contacts:
        contacts = [{
            'id': 1,
            'name': 'Contacto soporte',
            'email': support_email,
            'phone': '',
            'source': 'Configuración',
            'blocked': False
        }]

    if not profiles:
        profile_name = str(settings.get('perfil', {}).get('nombre_mostrar', '')).strip() or 'Perfil principal'
        profiles = [{
            'id': 1,
            'name': profile_name,
            'description': 'Visible para servicios internos'
        }]

    contacts_share['family'] = family
    contacts_share['preferences'] = preferences
    contacts_share['contacts'] = contacts
    contacts_share['profiles'] = profiles
    payload['contacts_share'] = contacts_share
    return contacts_share


def _contacts_share_summary(contacts_share):
    contacts = contacts_share.get('contacts', []) if isinstance(contacts_share, dict) else []
    profiles = contacts_share.get('profiles', []) if isinstance(contacts_share, dict) else []
    family = contacts_share.get('family', {}) if isinstance(contacts_share, dict) else {}
    preferences = contacts_share.get('preferences', {}) if isinstance(contacts_share, dict) else {}

    blocked_count = len([item for item in contacts if bool(item.get('blocked', False))])
    active_count = len([item for item in contacts if not bool(item.get('blocked', False))])

    return {
        'contacts_total': len(contacts),
        'contacts_active': active_count,
        'contacts_blocked': blocked_count,
        'profiles_total': len(profiles),
        'invites_remaining': max(_sanitize_int(family.get('invites_remaining', 0), 0), 0),
        'family_enabled': bool(family.get('enabled', False)),
        'share_location': bool(preferences.get('share_location', False))
    }


def _ensure_privacy_center(payload):
    privacy = payload.get('privacy_center', {})
    if not isinstance(privacy, dict):
        privacy = {}

    defaults = {
        'search_personalization': True,
        'play_personalization': True,
        'web_activity': True,
        'play_history': True,
        'youtube_history': True,
        'maps_timeline': False,
        'ads_personalized': True,
        'partner_ads': False,
        'legacy_plan_created': False,
        'delete_account_requested_at': '',
        'service_emails_enabled': True,
        'fit_data_enabled': True,
        'voice_match_enabled': False,
        'download_requests': 0
    }

    for key, default_value in defaults.items():
        if key not in privacy:
            privacy[key] = default_value

    privacy['download_requests'] = max(_sanitize_int(privacy.get('download_requests', 0), 0), 0)
    payload['privacy_center'] = privacy
    return privacy


def _privacy_center_summary(privacy, activity):
    active_controls = 0
    control_keys = [
        'search_personalization',
        'play_personalization',
        'web_activity',
        'play_history',
        'youtube_history',
        'ads_personalized',
        'service_emails_enabled',
        'fit_data_enabled',
        'voice_match_enabled'
    ]
    for key in control_keys:
        if bool(privacy.get(key, False)):
            active_controls += 1

    privacy_keywords = ['privacidad', 'actividad', 'anuncio', 'historial', 'datos']
    privacy_activity_count = 0
    for item in activity:
        haystack = f"{str(item.get('action', ''))} {str(item.get('detail', ''))}".lower()
        if any(keyword in haystack for keyword in privacy_keywords):
            privacy_activity_count += 1

    return {
        'active_controls': active_controls,
        'privacy_activity_count': privacy_activity_count,
        'download_requests': max(_sanitize_int(privacy.get('download_requests', 0), 0), 0),
        'legacy_plan_created': bool(privacy.get('legacy_plan_created', False)),
        'delete_account_requested_at': str(privacy.get('delete_account_requested_at', '')).strip()
    }

def extract_emails_from_df(df):
    emails = set()
    email_regex = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")
    for col in df.columns:
        for val in df[col].dropna():
            for m in email_regex.findall(str(val)):
                emails.add(m)
    return sorted(emails)

def extract_emails_without_voucher(df):
    if df is None or len(df) == 0:
        return []

    candidate_cols = []
    for col in df.columns:
        normalized = str(col).strip().lower().replace('°', 'º')
        if normalized in ['nº documento', 'nºdocumento', 'asientos', 'voucher contable']:
            candidate_cols.append(col)
        elif 'documento' in normalized or 'voucher' in normalized or 'asiento' in normalized:
            candidate_cols.append(col)

    if candidate_cols:
        voucher_col = candidate_cols[0]
        voucher_values = df[voucher_col]
        empty_mask = voucher_values.isna() | (voucher_values.astype(str).str.strip() == '')
        filtered_df = df.loc[empty_mask]
    else:
        filtered_df = df

    return extract_emails_from_df(filtered_df)

def normalize_reference(value):
    return str(value).strip().upper()

def extract_single_email(value):
    if pd.isna(value):
        return ''
    matches = re.findall(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}", str(value))
    if matches:
        return matches[0].strip().lower()
    return ''

def build_clientes_email_map(df_clientes):
    if df_clientes is None or len(df_clientes) == 0:
        return {}

    normalized_columns = [str(c).strip().lower() for c in df_clientes.columns]
    if 'referencia' not in normalized_columns or 'correo de contacto' not in normalized_columns:
        return {}

    referencia_col = df_clientes.columns[normalized_columns.index('referencia')]
    correo_col = df_clientes.columns[normalized_columns.index('correo de contacto')]

    df_map = df_clientes[[referencia_col, correo_col]].copy()
    df_map['referencia_key'] = df_map[referencia_col].apply(normalize_reference)
    df_map['correo_clean'] = df_map[correo_col].apply(extract_single_email)
    df_map = df_map[df_map['correo_clean'] != '']

    return dict(zip(df_map['referencia_key'], df_map['correo_clean']))

def load_clientes_email_map_from_bd():
    try:
        config_path = bd_path('config.json')
        if not os.path.exists(config_path):
            return {}
        import json
        with open(config_path, 'r', encoding='utf-8') as file:
            config = json.load(file)
        clientes_file_name = config.get('CLIENTES')
        if not clientes_file_name:
            return {}
        clientes_path = bd_path(clientes_file_name)
        if not os.path.exists(clientes_path):
            return {}
        df_clientes = pd.read_excel(clientes_path, header=0)
        return build_clientes_email_map(df_clientes)
    except Exception:
        return {}

def get_no_voucher_mask(df):
    if df is None or len(df) == 0:
        return pd.Series(dtype=bool)

    candidate_cols = []
    for col in df.columns:
        normalized = str(col).strip().lower().replace('°', 'º')
        if normalized in ['nº documento', 'nºdocumento', 'asientos', 'voucher contable']:
            candidate_cols.append(col)
        elif 'documento' in normalized or 'voucher' in normalized or 'asiento' in normalized:
            candidate_cols.append(col)

    if candidate_cols:
        voucher_col = candidate_cols[0]
        voucher_values = df[voucher_col]
        return voucher_values.isna() | (voucher_values.astype(str).str.strip() == '')

    return pd.Series([True] * len(df), index=df.index)

def collect_emails_without_voucher_using_clientes(df_movimientos, clientes_email_map):
    if df_movimientos is None or len(df_movimientos) == 0 or not clientes_email_map:
        return []

    no_voucher_mask = get_no_voucher_mask(df_movimientos)
    if len(no_voucher_mask) == 0:
        return []

    if 'Correo' not in df_movimientos.columns:
        df_movimientos['Correo'] = ''

    if 'Referencia' in df_movimientos.columns:
        for index, row in df_movimientos.loc[no_voucher_mask].iterrows():
            ref_key = normalize_reference(row.get('Referencia', ''))
            if ref_key in clientes_email_map:
                df_movimientos.at[index, 'Correo'] = clientes_email_map[ref_key]

    emails = set()
    for value in df_movimientos.loc[no_voucher_mask, 'Correo'].dropna():
        clean_email = extract_single_email(value)
        if clean_email:
            emails.add(clean_email)
    return sorted(emails)


def build_voucher_data_without_pdf(df_movimientos, clientes_email_map=None):
    if df_movimientos is None or len(df_movimientos) == 0:
        return []

    no_voucher_mask = get_no_voucher_mask(df_movimientos)
    if len(no_voucher_mask) == 0:
        return []

    voucher_records_by_email = {}

    for index, row in df_movimientos.loc[no_voucher_mask].iterrows():
        referencia = str(row.get('Referencia', row.get('referencia', f'REF-{index}'))).strip()

        email = ''
        if 'Correo' in row and pd.notna(row['Correo']) and str(row['Correo']).strip():
            email = extract_single_email(row['Correo'])
        elif 'Correos' in row and pd.notna(row['Correos']) and str(row['Correos']).strip():
            email = extract_single_email(str(row['Correos']).split(',')[0])
        elif clientes_email_map and referencia.upper() in clientes_email_map:
            email = extract_single_email(clientes_email_map[referencia.upper()])

        if not email:
            continue

        nombre_cliente = ''
        for candidate_col in [
            'Nombre Cliente', 'Nombre', 'Cliente', 'Razón Social', 'Razon Social',
            'Titular', 'Nombre y Apellido'
        ]:
            if candidate_col in row and pd.notna(row[candidate_col]) and str(row[candidate_col]).strip():
                nombre_cliente = str(row[candidate_col]).strip()
                break

        if not nombre_cliente:
            local_part = email.split('@')[0]
            name_tokens = [token for token in re.split(r'[._\-]+', local_part) if token and not token.isdigit()]
            if name_tokens:
                nombre_cliente = ' '.join(token.capitalize() for token in name_tokens[:3])

        if not nombre_cliente:
            nombre_cliente = 'Cliente'

        fecha_value = row.get('Fecha', row.get('fecha', ''))
        if isinstance(fecha_value, pd.Timestamp):
            fecha_value = fecha_value.strftime('%d/%m/%Y')

        voucher_records_by_email[email] = {
            'email': email,
            'referencia': referencia,
            'monto': row.get('Monto', row.get('monto', 0)),
            'nombre_cliente': nombre_cliente,
            'fecha': fecha_value,
            'descripcion': row.get('Descripción operación', row.get('descripcion', 'Operación bancaria')),
            'operacion_numero': row.get('Operación - Número', row.get('operacion', 'N/A')),
            'filepath': ''
        }

    return list(voucher_records_by_email.values())


def is_company_name(nombre_cliente):
    nombre = str(nombre_cliente or '').strip()
    if not nombre:
        return False

    nombre_upper = nombre.upper()
    has_company_word = any(word in nombre_upper for word in COMPANY_KEYWORDS)
    has_digits = bool(re.search(r'\d', nombre_upper))
    return has_company_word or has_digits


def build_saludo_cliente(nombre_cliente):
    nombre = str(nombre_cliente or '').strip()
    if not nombre or nombre.lower() == 'cliente':
        return 'Estimado Cliente,'
    if is_company_name(nombre):
        return 'Estimado Cliente,'
    return f'Estimado {nombre},'


def build_voucher_email_text(saludo):
    return f"""{saludo}

Le informamos que hemos recibido un abono en nuestra cuenta corriente del BANCO DE CREDITO con los siguientes detalle:



Para proceder con sus recibos, le invitamos a acceder a nuestra plataforma de Oficina Virtual ENSA.
https://servicios.distriluz.com.pe/oficinavirtual

En esta plataforma, podrá registrarse como Cliente Empresa para gestionar la cancelación de los suministros afiliados a su representada y agregar otros suministros. Podrá adjuntar la constancia del pago o transferencia realizada para completar el proceso.

Agradecemos su atención y quedamos a su disposición para cualquier consulta adicional al CEL. 979 450 731 o al correo electrónico: recaudacionensa@distriluz.com.pe
"""


def build_voucher_email_html(saludo, voucher_info=None):
    referencia = str((voucher_info or {}).get('referencia', '') or 'N/A')
    fecha = str((voucher_info or {}).get('fecha', '') or 'N/A')
    operacion = str((voucher_info or {}).get('operacion_numero', '') or 'N/A')
    descripcion = str((voucher_info or {}).get('descripcion', '') or 'Operación bancaria')
    monto = (voucher_info or {}).get('monto', 0)

    try:
        if isinstance(monto, str):
            monto_clean = monto.replace(',', '').strip()
            monto_value = float(monto_clean) if monto_clean else 0.0
        else:
            monto_value = float(monto)
        monto_text = f"S/ {monto_value:,.2f}"
    except Exception:
        monto_text = str(monto or 'S/ 0.00')

    return f"""
<html>
    <body style="margin:0; padding:0; background:#f3f6fb; font-family: 'Segoe UI', Arial, sans-serif; color:#1f2937; line-height:1.55;">
        <div style="max-width:760px; margin:20px auto; background:#ffffff; border:1px solid #e5e7eb; border-radius:12px; overflow:hidden;">
            <div style="background:linear-gradient(135deg, #1e5da8 0%, #274b8f 100%); color:#ffffff; padding:18px 22px;">
                <div style="font-size:13px; opacity:0.95; letter-spacing:0.2px;">ENSA</div>
                <div style="font-size:22px; font-weight:700; margin-top:2px;">Voucher de Abono Recibido</div>
            </div>

            <div style="padding:20px 22px 10px 22px;">
                <p style="margin:0 0 10px 0; font-size:18px; font-weight:700; color:#1e5da8;">
                    ENSA - Voucher de Abono Recibido
                </p>
                <p style="margin:0 0 14px 0; font-size:15px;">{saludo}</p>
                <p style="margin:0 0 14px 0; font-size:14px; color:#374151;">
                    Le informamos que hemos recibido un abono en nuestra cuenta corriente del BANCO DE CREDITO con los siguientes detalle:
                </p>

                <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="border:1px solid #dbe6f5; border-radius:10px; margin:0 0 14px 0; overflow:hidden;">
                    <tr>
                        <td style="background:#eef4fd; padding:10px 14px; font-size:13px; font-weight:700; color:#1e3a8a;">Detalle del abono</td>
                    </tr>
                    <tr>
                        <td style="padding:0; background:#ffffff;">
                            <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="font-size:14px; color:#1f2937;">
                                <tr>
                                    <td style="width:185px; padding:10px 14px; border-top:1px solid #e5edf9; color:#4b5563;"><strong>Referencia</strong></td>
                                    <td style="padding:10px 14px; border-top:1px solid #e5edf9;">{referencia}</td>
                                </tr>
                                <tr>
                                    <td style="width:185px; padding:10px 14px; border-top:1px solid #e5edf9; background:#f9fbff; color:#4b5563;"><strong>Fecha</strong></td>
                                    <td style="padding:10px 14px; border-top:1px solid #e5edf9; background:#f9fbff;">{fecha}</td>
                                </tr>
                                <tr>
                                    <td style="width:185px; padding:10px 14px; border-top:1px solid #e5edf9; color:#4b5563;"><strong>Número de operación</strong></td>
                                    <td style="padding:10px 14px; border-top:1px solid #e5edf9;">{operacion}</td>
                                </tr>
                                <tr>
                                    <td style="width:185px; padding:10px 14px; border-top:1px solid #e5edf9; background:#f9fbff; color:#4b5563;"><strong>Descripción</strong></td>
                                    <td style="padding:10px 14px; border-top:1px solid #e5edf9; background:#f9fbff;">{descripcion}</td>
                                </tr>
                                <tr>
                                    <td style="width:185px; padding:12px 14px; border-top:1px solid #d7e3f7; background:#eef4fd; color:#1e3a8a;"><strong>Monto</strong></td>
                                    <td style="padding:12px 14px; border-top:1px solid #d7e3f7; background:#eef4fd; font-size:17px; font-weight:700; color:#1e5da8;">{monto_text}</td>
                                </tr>
                            </table>
                        </td>
                    </tr>
                </table>

                <p style="margin:0 0 10px 0; font-size:14px;">
                    Para proceder con sus recibos, le invitamos a acceder a nuestra plataforma de <strong style="color:#1e5da8;">Oficina Virtual ENSA</strong>.
                </p>

                <p style="margin:0 0 14px 0; font-size:14px;">
                    <a href="https://servicios.distriluz.com.pe/oficinavirtual" style="color:#1e5da8; text-decoration:underline; font-weight:600;">
                        https://servicios.distriluz.com.pe/oficinavirtual
                    </a>
                </p>

                <p style="margin:0 0 12px 0; font-size:14px; color:#374151;">
                    En esta plataforma, podrá registrarse como Cliente Empresa para gestionar la cancelación de los suministros afiliados a su representada y agregar otros suministros.
                    <strong style="color:#1e5da8;">Podrá adjuntar la constancia del pago o transferencia realizada para completar el proceso.</strong>
                </p>
            </div>

            <div style="border-top:1px solid #e5e7eb; background:#fafbfc; padding:14px 22px; font-size:13px; color:#4b5563;">
                Agradecemos su atención y quedamos a su disposición para cualquier consulta adicional al CEL. 979 450 731 o al correo electrónico: <strong>recaudacionensa@distriluz.com.pe</strong>
            </div>
        </div>
    </body>
</html>
"""

def save_emails_cache(emails):
    try:
        cache_path = files_path('emails_cache.json')
        with open(cache_path, 'w', encoding='utf-8') as f:
            json.dump({'emails': emails}, f, ensure_ascii=False)
    except Exception:
        pass

def load_emails_cache():
    try:
        cache_path = files_path('emails_cache.json')
        if not os.path.exists(cache_path):
            return []
        with open(cache_path, 'r', encoding='utf-8') as f:
            payload = json.load(f)
        emails = payload.get('emails', [])
        if isinstance(emails, list):
            return emails
        return []
    except Exception:
        return []


def count_vouchers_in_folder():
    try:
        base_dir = vouchers_path()
        if not os.path.exists(base_dir):
            return 0
        total = 0
        for filename in os.listdir(base_dir):
            if filename.lower().endswith('.pdf'):
                total += 1
        return total
    except Exception:
        return 0


def find_latest_voucher_for_email(recipient_email):
    recipient = str(recipient_email or '').strip().lower()
    if not recipient:
        return None

    encoded_email = recipient.replace('@', '_').replace('.', '_')
    expected_suffix = f"_{encoded_email}.pdf"

    try:
        base_dir = vouchers_path()
        if not os.path.exists(base_dir):
            return None

        candidate_paths = []
        for filename in os.listdir(base_dir):
            filename_lower = filename.lower()
            if not filename_lower.endswith('.pdf'):
                continue
            if filename_lower.endswith(expected_suffix):
                candidate_paths.append(os.path.join(base_dir, filename))

        if not candidate_paths:
            return None

        latest_path = max(candidate_paths, key=os.path.getmtime)
        return {
            'email': recipient,
            'referencia': None,
            'filepath': latest_path,
            'monto': None
        }
    except Exception:
        return None

def render_correos_page(emails=None, mensaje_exito=None, page=1, quick_password_message=None, force_quick_password=False):
    if emails is None:
        emails = []
    page_size = 50
    total = len(emails)
    total_pages = max(1, (total + page_size - 1) // page_size)
    if page < 1:
        page = 1
    if page > total_pages:
        page = total_pages
    start = (page - 1) * page_size
    end = start + page_size
    page_emails = emails[start:end]
    
    # Obtener información de vouchers disponibles
    vouchers_generados = session.get('vouchers_generados', [])
    total_vouchers = len(vouchers_generados)
    secure_smtp = load_secure_smtp_credentials()
    smtp_config = {
        'sender': secure_smtp.get('sender', ''),
        'smtp_host': secure_smtp.get('smtp_host', '') or 'owa.fonafe.gob.pe',
        'smtp_port': secure_smtp.get('smtp_port', '') or '587',
        'smtp_security': secure_smtp.get('smtp_security', '') or 'starttls'
    }
    google_config = load_google_config()
    general_settings = load_general_settings()
    account_features = load_account_features()

    if not account_features.get('third_party_apps'):
        account_features['third_party_apps'] = _default_third_party_apps()

    contacts_share = _ensure_contacts_share(account_features, general_settings)
    privacy_center = _ensure_privacy_center(account_features)
    save_account_features(account_features)

    feature_security = account_features.get('security', {})
    sessions_list = account_features.get('sessions', [])
    active_sessions = len([item for item in sessions_list if bool(item.get('active', False))])
    third_party_total = len(account_features.get('third_party_apps', []))

    security_panel = {
        'recovery_phone': str(feature_security.get('recovery_phone', '')).strip(),
        'recovery_email': str(feature_security.get('recovery_email', '')).strip(),
        'two_factor_phone': str(feature_security.get('two_factor_phone', '')).strip(),
        'backup_codes': max(_sanitize_int(feature_security.get('backup_codes', 0), 0), 0),
        'device_prompt_count': max(_sanitize_int(feature_security.get('device_prompt_count', feature_security.get('google_prompt_devices', 0)), 0), 0),
        'skip_password_when_possible': bool(feature_security.get('skip_password_when_possible', False)),
        'enhanced_browsing': bool(feature_security.get('enhanced_browsing', False)),
        'sessions_total': len(sessions_list),
        'active_sessions': active_sessions,
        'third_party_total': third_party_total,
        'password_updated_at': get_account_password_updated_at()
    }

    contacts_panel = {
        'correo_soporte': str(general_settings.get('contactos', {}).get('correo_soporte', '')).strip(),
        'correo_copia': str(general_settings.get('contactos', {}).get('correo_copia', '')).strip(),
        'sync_interactions': bool(contacts_share.get('preferences', {}).get('sync_interactions', True)),
        'sync_device_contacts': bool(contacts_share.get('preferences', {}).get('sync_device_contacts', True)),
        'share_location': bool(contacts_share.get('preferences', {}).get('share_location', False)),
        'profile_visible': bool(contacts_share.get('preferences', {}).get('profile_visible', True)),
        'family_enabled': bool(contacts_share.get('family', {}).get('enabled', False)),
        'group_name': str(contacts_share.get('family', {}).get('group_name', 'Mi grupo familiar')).strip() or 'Mi grupo familiar',
        'summary': _contacts_share_summary(contacts_share)
    }

    privacy_panel = {
        'search_personalization': bool(privacy_center.get('search_personalization', True)),
        'play_personalization': bool(privacy_center.get('play_personalization', True)),
        'web_activity': bool(privacy_center.get('web_activity', True)),
        'play_history': bool(privacy_center.get('play_history', True)),
        'youtube_history': bool(privacy_center.get('youtube_history', True)),
        'maps_timeline': bool(privacy_center.get('maps_timeline', False)),
        'ads_personalized': bool(privacy_center.get('ads_personalized', True)),
        'partner_ads': bool(privacy_center.get('partner_ads', False)),
        'service_emails_enabled': bool(privacy_center.get('service_emails_enabled', True)),
        'fit_data_enabled': bool(privacy_center.get('fit_data_enabled', True)),
        'voice_match_enabled': bool(privacy_center.get('voice_match_enabled', False)),
        'summary': _privacy_center_summary(privacy_center, account_features.get('activity', []))
    }

    photo_path = _profile_photo_path()
    has_profile_photo = os.path.exists(photo_path)
    profile_photo_version = general_settings.get('perfil', {}).get('foto_version', 0)
    
    return render_template(
        'correos.html',
        emails=emails,
        page_emails=page_emails,
        page=page,
        total_pages=total_pages,
        page_size=page_size,
        total_emails=total,
        mensaje_exito=mensaje_exito,
        quick_password_message=quick_password_message,
        force_quick_password=force_quick_password,
        total_vouchers=total_vouchers,
        vouchers_generados=vouchers_generados,
        smtp_config=smtp_config,
        google_config=google_config,
        general_settings=general_settings,
        security_panel=security_panel,
        contacts_panel=contacts_panel,
        privacy_panel=privacy_panel,
        has_profile_photo=has_profile_photo,
        profile_photo_url=f"/foto_perfil_actual?v={profile_photo_version}"
    )

def extract_emails_from_excel_upload(file_storage):
    emails = set()
    email_regex = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")
    try:
        file_storage.stream.seek(0)
        workbook = load_workbook(filename=BytesIO(file_storage.read()), data_only=True)
        file_storage.stream.seek(0)

        for ws in workbook.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        for m in email_regex.findall(str(cell.value)):
                            emails.add(m)
                    if cell.hyperlink and cell.hyperlink.target:
                        target = str(cell.hyperlink.target)
                        if target.lower().startswith('mailto:'):
                            target = target[7:]
                        target = target.split('?', 1)[0]
                        for m in email_regex.findall(target):
                            emails.add(m)
    except Exception:
        # If the file cannot be opened with openpyxl (e.g., unsupported format), fallback to df extraction only
        try:
            file_storage.stream.seek(0)
        except Exception:
            pass
    return sorted(emails)


def _enforce_step_flow(current_step: str):
    return None

@app.route('/', methods=['POST','GET'])
def home():
    auth_redirect = _require_worker_microsoft_login()
    if auth_redirect is not None:
        return auth_redirect

    flow_redirect = _enforce_step_flow('home')
    if flow_redirect is not None:
        return flow_redirect

    if request.method == 'GET' and request.args.get('reset') == '1':
        session.pop('home_processing_result', None)
        session.pop('home_success_message', None)

    general_settings = load_general_settings()
    photo_path = _profile_photo_path()
    has_profile_photo = os.path.exists(photo_path)
    profile_photo_version = general_settings.get('perfil', {}).get('foto_version', 0)

    if request.method == 'POST':
        files = request.files.getlist('file')
        filtered_files = [x for x in files if x.filename!=""]
        if len(filtered_files) <= 1:
            return render_template(
                'home.html',
                error_message='Debe subir por lo menos un archivo.',
                processing_result=session.get('home_processing_result'),
                mensaje_exito=session.get('home_success_message'),
                has_profile_photo=has_profile_photo,
                profile_photo_url=f"/foto_perfil_actual?v={profile_photo_version}"
            )
        else:
            try:
                accountService = AccountService()    
                transferService = TransferService()
                interbankService = InterbankService()
                providerService = ProviderService()
                for file in files:
                    nombre =  file.filename.upper() 
                    if (nombre != ""):
                        if CUENTA in nombre:
                            accountService.process_movements(file)
                        elif TRANSFER in nombre:  
                            transferService.setMovimientos(accountService.movimientos)
                            transferService.process_transfers(file)
                        elif INTERBANK in nombre:
                            interbankService.setMovimientos(accountService.movimientos)
                            interbankService.process_interbanks(file)
                        elif PROVIDERS in nombre:
                            providerService.setMovimientos(accountService.movimientos)
                            providerService.process_providers(file)    
                        else:
                            raise Exception("Archivo no ubicado: "+nombre)    
                guardaMovimientos(accountService.movimientos)
                guardaRecaudos(accountService.recaudos)
                resumen = {"movements": accountService.error, "providers": providerService.error, "transfers": transferService.error, "interbanks": interbankService.error}
                session['required_next_step'] = 'asiento'
                session['home_processing_result'] = resumen
                session['home_success_message'] = 'Proceso exitosamente.'
                return redirect(url_for('home'))
    
                #return redirect(url_for('upload'))
            except Exception as e:
                error_message = str(e)
                return render_template(
                    'home.html',
                    error_message=error_message,
                    processing_result=session.get('home_processing_result'),
                    mensaje_exito=session.get('home_success_message'),
                    has_profile_photo=has_profile_photo,
                    profile_photo_url=f"/foto_perfil_actual?v={profile_photo_version}"
                )
    else:
        return render_template(
            'home.html',
            processing_result=session.get('home_processing_result'),
            mensaje_exito=session.get('home_success_message'),
            has_profile_photo=has_profile_photo,
            profile_photo_url=f"/foto_perfil_actual?v={profile_photo_version}"
        )


@app.route('/menu', methods=['GET'])
def menu():
    auth_redirect = _require_worker_microsoft_login()
    if auth_redirect is not None:
        return auth_redirect

    tab = request.args.get('tab', 'home').strip().lower()
    tab_map = {
        'home': '/',
        'basedatos': '/basedatos',
        'asiento': '/asiento',
        'correos': '/correos'
    }
    if tab not in tab_map:
        tab = 'home'

    return render_template('menu.html', active_tab=tab, tab_map=tab_map)

def guardaMovimientos(movimientos):
    movimientos["Fecha"] = pd.to_datetime(movimientos["Fecha"], format="%d/%m/%Y").dt.strftime("%d/%m/%Y")
    excel_file = BytesIO()
    movimientos.to_excel(excel_file, index=False)
    excel_file.seek(0)
    workbook = openpyxl.load_workbook(excel_file)
    worksheet = workbook.active 
    worksheet.column_dimensions["A"].width = 20  
    worksheet.column_dimensions["C"].width = 30  
    worksheet.column_dimensions["K"].width = 40  
    worksheet.column_dimensions["L"].width = 40  
    worksheet.column_dimensions["M"].width = 35 
    worksheet.column_dimensions["N"].width = 40
    for col in worksheet.iter_cols(min_row=1, max_row=1):
        header_value = col[0].value
        if header_value == 'Correo':
            worksheet.column_dimensions[col[0].column_letter].width = 42
            break
    ruta_archivo = files_path('movimientos.xlsx')
    workbook.save(ruta_archivo)

def guardaRecaudos(recaudos):
    excel_file = BytesIO()
    recaudos.to_excel(excel_file, index=False)
    excel_file.seek(0)
    workbook = openpyxl.load_workbook(excel_file)
    worksheet = workbook.active 
    worksheet.column_dimensions["A"].width = 15  
    worksheet.column_dimensions["B"].width = 50  
    worksheet.column_dimensions["C"].width = 25  
    worksheet.column_dimensions["D"].width = 10  
    worksheet.column_dimensions["E"].width = 15 
    worksheet.column_dimensions["F"].width = 10 

    ruta_archivo = files_path('recaudos.xlsx')
    workbook.save(ruta_archivo)

@app.route('/basedatos', methods=['POST','GET'])
def basedatos():
    auth_redirect = _require_worker_microsoft_login()
    if auth_redirect is not None:
        return auth_redirect

    flow_redirect = _enforce_step_flow('basedatos')
    if flow_redirect is not None:
        return flow_redirect

    if request.method == 'POST':
        files    = request.files.getlist('file')
        
        try:
            filtered_files = [x for x in files if x.filename!=""]
                        
            if len(filtered_files) < 1:
                return render_template('base-datos.html', error_message= 'Debe subir por lo menos un archivo.')

            nombres = [f.filename.upper() for f in filtered_files]
            valid_patterns = ['RECAUDO', 'PREPAGO', 'TRABAJADOR', 'CLIENTE']
            invalid_files = [n for n in nombres if not any(pattern in n for pattern in valid_patterns)]
            if invalid_files:
                return render_template('base-datos.html', error_message='Archivo(s) no reconocido(s): ' + ', '.join(invalid_files))

            mensaje_exito = 'Proceso exitosamente.'
            
            base_datos_service = BaseDatosService()  
            base_datos_service.GuardarAchivos(files)  
            session['required_next_step'] = 'home'
            return render_template('base-datos.html', mensaje_exito=mensaje_exito)
                
        except Exception as e:
            error_message = str(e)
            return render_template('base-datos.html', error_message= error_message)

    else:
        nohay = 'Archivo subido correctamente.'
        return render_template('base-datos.html')
    
@app.route('/asiento', methods=['POST'])
def asiento_procesar():
    auth_redirect = _require_worker_microsoft_login()
    if auth_redirect is not None:
        return auth_redirect

    flow_redirect = _enforce_step_flow('asiento')
    if flow_redirect is not None:
        return flow_redirect

    logging.error('asiento_procesar: start')
    files = request.files.getlist('file')
    logging.error('asiento_procesar: received %d files', len(files))
    filtered_files = [x for x in files if x.filename!=""]
    logging.error('asiento_procesar: filtered %d files', len(filtered_files))
    if len(filtered_files) <= 1:
        logging.error('asiento_procesar: not enough files, returning form')
        return render_template('asiento.html', error_message= 'Debe subir ambos archivo.')
    else:
        try:
            asientoService = AsientoService()    
            # detect files: movimientos y asientos
            movimientosfile = None
            asientosfile = None
            for file in files:
                nombre =  file.filename.upper()
                if (nombre != ""):
                    if MOVIMIENTOS in nombre or "MOVIMIENTO" in nombre:
                        movimientosfile = file
                    elif ASIENTO in nombre:
                        asientosfile = file
                    else:
                        # ignore unknown files for now
                        pass

            if movimientosfile is None or asientosfile is None:
                raise Exception('Faltan archivos requeridos: Movimientos o Asientos')

            clientes_email_map = load_clientes_email_map_from_bd()

            asientoService.conciliar(movimientosfile, asientosfile)
            #solo si hay asientos se completa en el cache
            if asientoService.df_movimientos is not None:
                emails = collect_emails_without_voucher_using_clientes(asientoService.df_movimientos, clientes_email_map)
                guardaAsientos(asientoService.df_movimientos)
                
                # PREPARAR INFORMACIÓN DE VOUCHERS EN HTML (SIN GENERAR PDF)
                try:
                    vouchers_generados = build_voucher_data_without_pdf(
                        asientoService.df_movimientos, 
                        clientes_email_map
                    )
                    session['vouchers_generados'] = vouchers_generados
                    logging.info(f'Vouchers HTML preparados: {len(vouchers_generados)}')
                except Exception as ve:
                    logging.error(f'Error preparando vouchers HTML: {str(ve)}')
                    session['vouchers_generados'] = []
                
                # guardar en session para uso posterior y redirigir al flujo de correos
                sorted_emails = sorted(set(emails))
                session['asiento_emails'] = sorted_emails
                session['correos_ready'] = True
                save_emails_cache(sorted_emails)
                if len(sorted_emails) == 0:
                    session['asiento_email_warning'] = 'No se encontraron correos para líneas sin voucher. Verifique Referencia y CORREO DE CONTACTO en Base de Datos > Clientes.'
                else:
                    session.pop('asiento_email_warning', None)
                session['required_next_step'] = 'correos'
                # guardaAsientos ya escribió files/asientos.xlsx, descargamos directamente
                ruta_archivo = files_path('asientos.xlsx')
                return send_file(ruta_archivo, as_attachment=True, download_name='asientos.xlsx')
            else: 
                #si hubiera error se pinta la misma pagina y no se redirecciona
                return render_template('asiento.html', error_message= 'No se encontro ningun asiento en el proceso')       
            
        except Exception as e:
            error_message = str(e)
            logging.error('asiento_procesar: exception: %s', error_message)
            return render_template('asiento.html', error_message= error_message)
    # Fallback: ensure the view always returns a response
    logging.error('asiento_procesar: reached end of function without explicit return')
    return render_template('asiento.html', error_message='Error inesperado en el procesamiento')



@app.route('/asiento', methods=['GET'])
def asiento_get():
    auth_redirect = _require_worker_microsoft_login()
    if auth_redirect is not None:
        return auth_redirect

    flow_redirect = _enforce_step_flow('asiento')
    if flow_redirect is not None:
        return flow_redirect

    show_result_mode = request.args.get('resultado_correo', '0') == '1'
    asiento_emails = session.get('asiento_emails', [])
    vouchers_generados = session.get('vouchers_generados', [])

    return render_template(
        'asiento.html',
        show_result_mode=show_result_mode,
        asiento_emails=asiento_emails,
        total_vouchers=len(vouchers_generados)
    )


@app.route('/correos', methods=['GET','POST'])
def correos():
    auth_redirect = _require_worker_microsoft_login()
    if auth_redirect is not None:
        return auth_redirect

    flow_redirect = _enforce_step_flow('correos')
    if flow_redirect is not None:
        return flow_redirect

    arrived_from_asiento = session.get('required_next_step') == 'correos'
    if arrived_from_asiento:
        session.pop('required_next_step', None)

    ready_from_asiento = bool(session.pop('correos_ready', False))
    allow_results = arrived_from_asiento and ready_from_asiento

    if allow_results:
        sess_emails = session.get('asiento_emails', [])
    else:
        session.pop('asiento_emails', None)
        session.pop('vouchers_generados', None)
        session.pop('asiento_email_warning', None)
        sess_emails = []

    warning_message = session.pop('asiento_email_warning', None)
    config_message = session.pop('config_message', None)
    quick_password_message = session.pop('quick_password_message', None)
    force_quick_password = not bool(session.get('quick_password_verified', False))
    page_message = config_message or warning_message

    try:
        page = int(request.args.get('page', '1'))
    except ValueError:
        page = 1
    if page < 1:
        page = 1

    return render_correos_page(
        emails=sess_emails,
        mensaje_exito=page_message,
        page=page,
        quick_password_message=quick_password_message,
        force_quick_password=force_quick_password
    )


@app.route('/gestor_contrasenas', methods=['GET'])
def gestor_contrasenas():
    add_account_activity('Gestor de contraseñas', 'Apertura de panel')
    message = session.pop('password_manager_message', None)
    updated_at = get_account_password_updated_at()
    return render_template(
        'gestor_contrasenas.html',
        message=message,
        password_updated_at=updated_at
    )


@app.route('/gestor_contrasenas/cambiar', methods=['POST'])
def gestor_contrasenas_cambiar():
    current_password = request.form.get('current_password', '').strip()
    new_password = request.form.get('new_password', '').strip()
    confirm_password = request.form.get('confirm_password', '').strip()

    if len(new_password) < 6:
        session['password_manager_message'] = 'La nueva contraseña debe tener al menos 6 caracteres.'
        return redirect(url_for('gestor_contrasenas'))

    if new_password != confirm_password:
        session['password_manager_message'] = 'La confirmación no coincide con la nueva contraseña.'
        return redirect(url_for('gestor_contrasenas'))

    existing_password = get_account_password_value()
    if existing_password and current_password != existing_password:
        session['password_manager_message'] = 'La contraseña actual no es correcta.'
        return redirect(url_for('gestor_contrasenas'))

    try:
        save_account_password(new_password)
        add_account_activity('Gestor de contraseñas', 'Contraseña actualizada desde gestor')
        session['password_manager_message'] = '✅ Contraseña actualizada correctamente desde el gestor.'
    except Exception as ex:
        session['password_manager_message'] = f'Error actualizando contraseña: {ex}'

    return redirect(url_for('gestor_contrasenas'))


@app.route('/mi_contrasena', methods=['GET'])
def mi_contrasena():
    add_account_activity('Mi contraseña', 'Apertura de panel')
    message = session.pop('password_message', None)
    updated_at = get_account_password_updated_at()
    return render_template('mi_contrasena.html', password_updated_at=updated_at, message=message)


@app.route('/mi_contrasena/guardar', methods=['POST'])
def mi_contrasena_guardar():
    new_password = request.form.get('new_password', '').strip()
    confirm_password = request.form.get('confirm_password', '').strip()

    if len(new_password) < 6:
        session['password_message'] = 'La contraseña debe tener al menos 6 caracteres.'
        return redirect(url_for('mi_contrasena'))

    if new_password != confirm_password:
        session['password_message'] = 'La confirmación no coincide con la contraseña.'
        return redirect(url_for('mi_contrasena'))

    try:
        save_account_password(new_password)
        add_account_activity('Mi contraseña', 'Contraseña actualizada')
        session['password_message'] = '✅ Contraseña actualizada correctamente.'
    except Exception as ex:
        session['password_message'] = f'Error guardando contraseña: {ex}'

    return redirect(url_for('mi_contrasena'))


@app.route('/seguridad_inicio_sesion', methods=['GET'])
def seguridad_inicio_sesion():
    add_account_activity('Seguridad e inicio de sesión', 'Apertura de panel')

    settings = load_general_settings()
    payload = load_account_features()
    secure_smtp = load_secure_smtp_credentials()

    security_settings = settings.get('seguridad', {})
    feature_security = payload.get('security', {})
    sessions_list = payload.get('sessions', [])
    profile_data = settings.get('perfil', {})

    connected_sender = str(secure_smtp.get('sender', '')).strip() or get_connected_account_email()

    if not payload.get('third_party_apps'):
        payload['third_party_apps'] = _default_third_party_apps()
        save_account_features(payload)

    recovery_email = str(feature_security.get('recovery_email', '')).strip()
    recovery_phone = str(feature_security.get('recovery_phone', '')).strip()

    if not recovery_email:
        recovery_email = str(profile_data.get('correo_alterno', '')).strip() or connected_sender
    if not recovery_phone:
        recovery_phone = str(profile_data.get('telefono', '')).strip()

    security_view = {
        'two_factor_enabled': bool(security_settings.get('doble_factor', False)),
        'skip_password_when_possible': bool(feature_security.get('skip_password_when_possible', False)),
        'enhanced_browsing': bool(feature_security.get('enhanced_browsing', False)),
        'recovery_email': recovery_email,
        'recovery_phone': recovery_phone,
        'two_factor_phone': str(feature_security.get('two_factor_phone', '')).strip(),
        'backup_codes': max(_sanitize_int(feature_security.get('backup_codes', 0), 0), 0),
        'device_prompt_count': max(_sanitize_int(feature_security.get('device_prompt_count', feature_security.get('google_prompt_devices', 0)), 0), 0)
    }

    sessions_summary = _group_sessions_summary(sessions_list)
    devices_total = sum(sessions_summary.values())
    message = session.pop('security_message', None)

    return render_template(
        'seguridad_inicio_sesion.html',
        message=message,
        security=security_view,
        password_updated_at=get_account_password_updated_at(),
        security_activity=_security_activity_items(limit=8),
        sessions_summary=sessions_summary,
        devices_total=devices_total,
        third_party_apps=payload.get('third_party_apps', []),
        connected_sender=connected_sender
    )


@app.route('/seguridad_inicio_sesion/guardar', methods=['POST'])
def seguridad_inicio_sesion_guardar():
    settings = load_general_settings()
    payload = load_account_features()

    security_settings = settings.get('seguridad', {})
    feature_security = payload.get('security', {})

    two_factor_enabled = request.form.get('two_factor_enabled', '').strip().lower() == 'on'
    skip_password = request.form.get('skip_password_when_possible', '').strip().lower() == 'on'
    enhanced_browsing = request.form.get('enhanced_browsing', '').strip().lower() == 'on'

    recovery_phone = request.form.get('recovery_phone', '').strip()
    recovery_email = request.form.get('recovery_email', '').strip().lower()
    two_factor_phone = request.form.get('two_factor_phone', '').strip()
    backup_codes = max(_sanitize_int(request.form.get('backup_codes', '0'), 0), 0)
    device_prompt_count = max(_sanitize_int(request.form.get('device_prompt_count', request.form.get('google_prompt_devices', '0')), 0), 0)

    security_settings['doble_factor'] = two_factor_enabled
    settings['seguridad'] = security_settings

    feature_security['skip_password_when_possible'] = skip_password
    feature_security['enhanced_browsing'] = enhanced_browsing
    feature_security['recovery_phone'] = recovery_phone
    feature_security['recovery_email'] = recovery_email
    feature_security['two_factor_phone'] = two_factor_phone
    feature_security['backup_codes'] = backup_codes
    feature_security['device_prompt_count'] = device_prompt_count
    payload['security'] = feature_security

    try:
        save_general_settings(settings)
        save_account_features(payload)
        add_account_activity('Seguridad e inicio de sesión', 'Opciones de acceso y recuperación actualizadas')
        session['security_message'] = '✅ Configuración de seguridad guardada correctamente.'
    except Exception as ex:
        session['security_message'] = f'Error guardando configuración de seguridad: {ex}'

    return redirect(url_for('seguridad_inicio_sesion'))


@app.route('/seguridad_inicio_sesion/cerrar_otras_sesiones', methods=['POST'])
def seguridad_cerrar_otras_sesiones():
    payload = load_account_features()
    current_session_id = str(session.get('account_session_id', '')).strip()

    affected = _deactivate_other_sessions(payload, current_session_id)
    save_account_features(payload)

    add_account_activity('Seguridad e inicio de sesión', f'Otras sesiones cerradas: {affected}')
    session['security_message'] = f'✅ Se cerraron {affected} sesión(es) en otros dispositivos.'
    return redirect(url_for('seguridad_inicio_sesion'))


@app.route('/seguridad_inicio_sesion/conexion/agregar', methods=['POST'])
def seguridad_conexion_agregar():
    app_name = request.form.get('app_name', '').strip()
    app_description = request.form.get('app_description', '').strip()

    if not app_name:
        session['security_message'] = 'Ingresa el nombre de la aplicación a conectar.'
        return redirect(url_for('seguridad_inicio_sesion'))

    payload = load_account_features()
    third_party_apps = payload.get('third_party_apps', [])
    next_id = max([_sanitize_int(item.get('id', 0), 0) for item in third_party_apps], default=0) + 1

    third_party_apps.append({
        'id': next_id,
        'name': app_name,
        'description': app_description or 'Integración personalizada',
        'connected_at': pd.Timestamp.now().strftime('%d/%m/%Y %H:%M:%S')
    })

    payload['third_party_apps'] = third_party_apps
    save_account_features(payload)

    add_account_activity('Seguridad e inicio de sesión', f'Conexión de tercero agregada: {app_name}')
    session['security_message'] = '✅ Aplicación conectada correctamente.'
    return redirect(url_for('seguridad_inicio_sesion'))


@app.route('/seguridad_inicio_sesion/conexion/eliminar/<int:app_id>', methods=['POST'])
def seguridad_conexion_eliminar(app_id):
    payload = load_account_features()
    third_party_apps = payload.get('third_party_apps', [])
    kept = [item for item in third_party_apps if _sanitize_int(item.get('id', 0), 0) != int(app_id)]

    if len(kept) == len(third_party_apps):
        session['security_message'] = 'No se encontró la conexión seleccionada.'
        return redirect(url_for('seguridad_inicio_sesion'))

    payload['third_party_apps'] = kept
    save_account_features(payload)

    add_account_activity('Seguridad e inicio de sesión', f'Conexión de tercero eliminada: id={app_id}')
    session['security_message'] = '✅ Conexión eliminada correctamente.'
    return redirect(url_for('seguridad_inicio_sesion'))


@app.route('/dispositivos', methods=['GET'])
def dispositivos():
    add_account_activity('Dispositivos', 'Apertura de panel')
    payload = load_account_features()
    current_session_id = session.get('account_session_id', '')

    grouped = {}
    sessions_list = payload.get('sessions', [])
    now_value = pd.Timestamp.now()
    for item in sessions_list:
        os_family = str(item.get('os_family', 'Otro')).strip() or 'Otro'
        grouped.setdefault(os_family, [])

        last_seen_text = str(item.get('last_seen', '')).strip()
        inactive_days = 0
        if last_seen_text:
            try:
                parsed_last_seen = pd.to_datetime(last_seen_text, format='%d/%m/%Y %H:%M:%S', errors='coerce')
                if pd.isna(parsed_last_seen):
                    raise ValueError('Fecha inválida')
                inactive_days = max((now_value - parsed_last_seen).days, 0)
            except Exception:
                inactive_days = 0

        grouped[os_family].append({
            'session_id': str(item.get('session_id', '')).strip(),
            'device_type': str(item.get('device_type', '')).strip() or 'Equipo',
            'ip': str(item.get('ip', '')).strip() or 'desconocido',
            'last_seen': _relative_time_label(last_seen_text),
            'user_agent': str(item.get('user_agent', '')).strip(),
            'active': bool(item.get('active', False)),
            'inactive_days': inactive_days,
            'is_current': str(item.get('session_id', '')).strip() == current_session_id
        })

    message = session.pop('devices_message', None)
    return render_template(
        'dispositivos.html',
        devices=payload.get('devices', []),
        sessions_grouped=grouped,
        total_sessions=len(sessions_list),
        message=message
    )


@app.route('/dispositivos/agregar', methods=['POST'])
def dispositivos_agregar():
    name = request.form.get('name', '').strip()
    location = request.form.get('location', '').strip()

    if not name:
        session['devices_message'] = 'Indica el nombre del dispositivo.'
        return redirect(url_for('dispositivos'))

    payload = load_account_features()
    devices = payload.get('devices', [])
    next_id = max([int(d.get('id', 0) or 0) for d in devices], default=0) + 1
    devices.append({
        'id': next_id,
        'name': name,
        'location': location or 'Sin ubicación',
        'active': True
    })
    payload['devices'] = devices
    save_account_features(payload)

    add_account_activity('Dispositivos', f'Dispositivo agregado: {name}')
    session['devices_message'] = '✅ Dispositivo agregado correctamente.'
    return redirect(url_for('dispositivos'))


@app.route('/dispositivos/eliminar/<int:device_id>', methods=['POST'])
def dispositivos_eliminar(device_id):
    payload = load_account_features()
    devices = payload.get('devices', [])
    kept = [d for d in devices if int(d.get('id', 0) or 0) != int(device_id)]

    if len(kept) == len(devices):
        session['devices_message'] = 'No se encontró el dispositivo seleccionado.'
        return redirect(url_for('dispositivos'))

    payload['devices'] = kept
    save_account_features(payload)
    add_account_activity('Dispositivos', f'Dispositivo eliminado: id={device_id}')
    session['devices_message'] = '✅ Dispositivo eliminado correctamente.'
    return redirect(url_for('dispositivos'))


@app.route('/dispositivos/desconectar/<session_id_value>', methods=['POST'])
def dispositivos_desconectar(session_id_value):
    payload = load_account_features()
    sessions_list = payload.get('sessions', [])

    target_session = str(session_id_value or '').strip()
    kept_sessions = [
        item for item in sessions_list
        if str(item.get('session_id', '')).strip() != target_session
    ]

    if len(kept_sessions) != len(sessions_list):
        payload['sessions'] = kept_sessions
        save_account_features(payload)
        add_account_activity('Dispositivos', f'Sesión desconectada: {session_id_value}')
        session['devices_message'] = '✅ Sesión desconectada y removida de la lista.'
    else:
        session['devices_message'] = 'No se encontró la sesión seleccionada.'

    return redirect(url_for('dispositivos'))


@app.route('/mi_actividad', methods=['GET'])
def mi_actividad():
    payload = load_account_features()
    activity = payload.get('activity', [])

    query_text = request.args.get('q', '').strip().lower()
    action_filter = request.args.get('action', '').strip().lower()

    filtered = []
    for item in activity:
        action_value = str(item.get('action', '')).strip()
        detail_value = str(item.get('detail', '')).strip()
        timestamp_value = str(item.get('timestamp', '')).strip()

        haystack = f"{action_value} {detail_value} {timestamp_value}".lower()
        if query_text and query_text not in haystack:
            continue
        if action_filter and action_filter not in action_value.lower():
            continue
        filtered.append(item)

    action_types = sorted({str(item.get('action', '')).strip() for item in activity if str(item.get('action', '')).strip()})
    total = len(activity)
    total_filtered = len(filtered)

    return render_template(
        'mi_actividad.html',
        activity=filtered,
        total=total,
        total_filtered=total_filtered,
        action_types=action_types,
        current_query=query_text,
        current_action=action_filter
    )


@app.route('/mi_actividad/limpiar', methods=['POST'])
def mi_actividad_limpiar():
    mode = request.form.get('mode', '').strip().lower() or 'all'
    if mode == 'last50':
        clear_account_activity(mode='last', keep_last=50)
        add_account_activity('Mi actividad', 'Se conservaron solo 50 registros recientes')
    else:
        clear_account_activity(mode='all')
        add_account_activity('Mi actividad', 'Actividad eliminada por usuario')

    return redirect(url_for('mi_actividad'))


@app.route('/datos_privacidad', methods=['GET'])
def datos_privacidad():
    add_account_activity('Datos y privacidad', 'Apertura de panel')

    settings = load_general_settings()
    payload = load_account_features()
    privacy_center = _ensure_privacy_center(payload)
    contacts_share = _ensure_contacts_share(payload, settings)

    third_party_apps = payload.get('third_party_apps', [])
    if not third_party_apps:
        third_party_apps = _default_third_party_apps()
        payload['third_party_apps'] = third_party_apps

    save_account_features(payload)

    history_status = {
        'web_activity': 'Activada' if privacy_center.get('web_activity', True) else 'Detenida',
        'play_history': 'Activada' if privacy_center.get('play_history', True) else 'Detenida',
        'youtube_history': 'Activada' if privacy_center.get('youtube_history', True) else 'Detenida',
        'maps_timeline': 'Activada' if privacy_center.get('maps_timeline', False) else 'Detenida'
    }

    message = session.pop('privacy_message', None)
    return render_template(
        'datos_privacidad.html',
        message=message,
        privacy=privacy_center,
        history_status=history_status,
        privacy_summary=_privacy_center_summary(privacy_center, payload.get('activity', [])),
        third_party_apps=third_party_apps,
        third_party_count=len(third_party_apps),
        contacts_summary=_contacts_share_summary(contacts_share)
    )


@app.route('/datos_privacidad/guardar', methods=['POST'])
def datos_privacidad_guardar():
    settings = load_general_settings()
    payload = load_account_features()
    privacy_center = _ensure_privacy_center(payload)

    privacy_center['search_personalization'] = request.form.get('search_personalization', '').strip().lower() == 'on'
    privacy_center['play_personalization'] = request.form.get('play_personalization', '').strip().lower() == 'on'
    privacy_center['web_activity'] = request.form.get('web_activity', '').strip().lower() == 'on'
    privacy_center['play_history'] = request.form.get('play_history', '').strip().lower() == 'on'
    privacy_center['youtube_history'] = request.form.get('youtube_history', '').strip().lower() == 'on'
    privacy_center['maps_timeline'] = request.form.get('maps_timeline', '').strip().lower() == 'on'
    privacy_center['ads_personalized'] = request.form.get('ads_personalized', '').strip().lower() == 'on'
    privacy_center['partner_ads'] = request.form.get('partner_ads', '').strip().lower() == 'on'
    privacy_center['service_emails_enabled'] = request.form.get('service_emails_enabled', '').strip().lower() == 'on'
    privacy_center['fit_data_enabled'] = request.form.get('fit_data_enabled', '').strip().lower() == 'on'
    privacy_center['voice_match_enabled'] = request.form.get('voice_match_enabled', '').strip().lower() == 'on'

    payload['privacy_center'] = privacy_center

    settings['privacidad']['compartir_datos'] = bool(privacy_center['partner_ads'])
    settings['privacidad']['analytics'] = bool(privacy_center['web_activity'])

    try:
        save_account_features(payload)
        save_general_settings(settings)
        add_account_activity('Datos y privacidad', 'Preferencias de privacidad actualizadas')
        session['privacy_message'] = '✅ Datos y privacidad guardados correctamente.'
    except Exception as ex:
        session['privacy_message'] = f'Error guardando datos y privacidad: {ex}'

    return redirect(url_for('datos_privacidad'))


@app.route('/datos_privacidad/historial/limpiar', methods=['POST'])
def datos_privacidad_limpiar_historial():
    payload = load_account_features()
    privacy_center = _ensure_privacy_center(payload)

    history_target = request.form.get('history_target', '').strip().lower()
    if history_target in ['web_activity', 'play_history', 'youtube_history', 'maps_timeline']:
        privacy_center[history_target] = False
        payload['privacy_center'] = privacy_center
        save_account_features(payload)
        add_account_activity('Datos y privacidad', f'Historial desactivado: {history_target}')
        session['privacy_message'] = f'✅ Historial actualizado: {history_target} ahora está detenido.'
    else:
        session['privacy_message'] = 'No se reconoció el historial a modificar.'

    return redirect(url_for('datos_privacidad'))


@app.route('/datos_privacidad/historial/detener/<history_target>', methods=['GET'])
def datos_privacidad_detener_historial(history_target):
    payload = load_account_features()
    privacy_center = _ensure_privacy_center(payload)

    target = str(history_target or '').strip().lower()
    if target in ['web_activity', 'play_history', 'youtube_history', 'maps_timeline']:
        privacy_center[target] = False
        payload['privacy_center'] = privacy_center
        save_account_features(payload)
        add_account_activity('Datos y privacidad', f'Historial detenido desde acceso rápido: {target}')
        session['privacy_message'] = f'✅ Historial detenido: {target}.'
    else:
        session['privacy_message'] = 'No se reconoció el historial seleccionado.'

    return redirect(url_for('datos_privacidad'))


@app.route('/datos_privacidad/solicitar_descarga', methods=['POST'])
def datos_privacidad_solicitar_descarga():
    payload = load_account_features()
    privacy_center = _ensure_privacy_center(payload)
    privacy_center['download_requests'] = max(_sanitize_int(privacy_center.get('download_requests', 0), 0), 0) + 1
    payload['privacy_center'] = privacy_center
    save_account_features(payload)

    add_account_activity('Datos y privacidad', 'Solicitud de descarga de datos generada')
    session['privacy_message'] = '✅ Solicitud de descarga registrada correctamente.'
    return redirect(url_for('datos_privacidad'))


@app.route('/datos_privacidad/crear_legado', methods=['POST'])
def datos_privacidad_crear_legado():
    payload = load_account_features()
    privacy_center = _ensure_privacy_center(payload)
    privacy_center['legacy_plan_created'] = True
    payload['privacy_center'] = privacy_center
    save_account_features(payload)

    add_account_activity('Datos y privacidad', 'Plan de legado digital configurado')
    session['privacy_message'] = '✅ Plan de legado digital configurado.'
    return redirect(url_for('datos_privacidad'))


@app.route('/datos_privacidad/solicitar_eliminar_cuenta', methods=['POST'])
def datos_privacidad_solicitar_eliminar_cuenta():
    payload = load_account_features()
    privacy_center = _ensure_privacy_center(payload)
    privacy_center['delete_account_requested_at'] = pd.Timestamp.now().strftime('%d/%m/%Y %H:%M:%S')
    payload['privacy_center'] = privacy_center
    save_account_features(payload)

    add_account_activity('Datos y privacidad', 'Solicitud de eliminación de cuenta registrada')
    session['privacy_message'] = '✅ Solicitud de eliminación registrada. Requiere revisión administrativa.'
    return redirect(url_for('datos_privacidad'))


@app.route('/informacion_personal', methods=['GET'])
def informacion_personal():
    add_account_activity('Información personal', 'Apertura de panel')
    settings = load_general_settings()
    perfil = settings.get('perfil', {})
    message = session.pop('personal_info_message', None)
    photo_version = int(perfil.get('foto_version', 0) or 0)
    edit_field = request.args.get('edit', '').strip().lower()

    return render_template(
        'informacion_personal.html',
        perfil=perfil,
        sender_email=get_connected_account_email(),
        has_profile_photo=os.path.exists(_profile_photo_path()),
        profile_photo_url=f"/foto_perfil_actual?v={photo_version}",
        message=message,
        edit_field=edit_field
    )


@app.route('/informacion_personal/guardar', methods=['POST'])
def informacion_personal_guardar():
    settings = load_general_settings()
    perfil = settings.get('perfil', {})

    def update_field(field_name, default_value=''):
        if field_name in request.form:
            return request.form.get(field_name, '').strip()
        return str(perfil.get(field_name, default_value)).strip()

    perfil['nombre_mostrar'] = update_field('nombre_mostrar')
    perfil['genero'] = update_field('genero')
    perfil['correo_personal'] = update_field('correo_personal')
    perfil['correo_alterno'] = update_field('correo_alterno')
    perfil['telefono'] = update_field('telefono')
    perfil['fecha_nacimiento'] = update_field('fecha_nacimiento')
    perfil['idioma'] = update_field('idioma', 'Español (España)') or 'Español (España)'
    perfil['direccion_casa'] = update_field('direccion_casa')
    perfil['direccion_trabajo'] = update_field('direccion_trabajo')
    perfil['otras_direcciones'] = update_field('otras_direcciones')

    settings['perfil'] = perfil

    try:
        save_general_settings(settings)
        add_account_activity('Información personal', 'Datos personales actualizados')
        session['personal_info_message'] = '✅ Información personal guardada correctamente.'
    except Exception as ex:
        session['personal_info_message'] = f'Error guardando información personal: {ex}'

    return redirect(url_for('informacion_personal'))


@app.route('/correo_electronico', methods=['GET'])
def correo_electronico():
    auth_redirect = _require_worker_microsoft_login()
    if auth_redirect is not None:
        return auth_redirect

    add_account_activity('Correo electrónico', 'Apertura de panel')
    secure_smtp = load_secure_smtp_credentials()
    message = session.pop('email_settings_message', None)
    session_sender = str(session.get('worker_sender', '')).strip()
    session_smtp_host = str(session.get('worker_smtp_host', '')).strip()
    session_smtp_port = str(session.get('worker_smtp_port', '')).strip()
    session_smtp_security = str(session.get('worker_smtp_security', '')).strip().lower()
    linked_microsoft = bool(session.get('smtp_link_verified', False) and session_sender)
    auth_method = str(session.get('worker_auth_method', '')).strip() or (
        'API Microsoft Graph' if session.get('worker_login_via_graph', False) else 'SMTP OWA'
    )
    auth_timestamp = str(session.get('worker_login_at', '')).strip()
    settings = load_general_settings()
    perfil = settings.get('perfil', {}) if isinstance(settings, dict) else {}
    profile_display_name = str(perfil.get('nombre_mostrar', '')).strip() or _build_display_name_from_email(session_sender)
    linked_profile = {
        'display_name': profile_display_name,
        'sender': session_sender,
        'role': 'Administrador de su cuenta',
        'status': 'Vinculado y verificado' if linked_microsoft else 'Autenticado - verificación pendiente',
        'linked_at': str(session.get('worker_login_at', '')).strip()
    }
    return render_template(
        'correo_electronico.html',
        sender=session_sender or secure_smtp.get('sender', '') or get_connected_account_email(),
        smtp_host=session_smtp_host or secure_smtp.get('smtp_host', '') or 'owa.fonafe.gob.pe',
        smtp_port=session_smtp_port or secure_smtp.get('smtp_port', '') or '587',
        smtp_security=session_smtp_security or secure_smtp.get('smtp_security', '') or 'starttls',
        message=message,
        linked_microsoft=linked_microsoft,
        linked_sender=session_sender,
        linked_profile=linked_profile,
        auth_proof={
            'verified': linked_microsoft,
            'method': auth_method,
            'timestamp': auth_timestamp
        }
    )


@app.route('/iniciar_sesion', methods=['GET'])
def iniciar_sesion():
    return redirect(url_for('basedatos'))


@app.route('/iniciar_sesion', methods=['POST'])
@app.route('/correo_electronico/guardar', methods=['POST'])
def correo_electronico_guardar():
    is_login_route = request.path == '/iniciar_sesion'
    if is_login_route:
        is_blocked, lock_minutes = _get_login_lock_state()
        if is_blocked:
            session['login_message'] = (
                f'Has agotado los intentos permitidos. '
                f'Intenta nuevamente en {lock_minutes} minuto(s).'
            )
            return redirect(url_for('iniciar_sesion'))

    existing = load_secure_smtp_credentials()
    sender = request.form.get('sender', '').strip()
    password = request.form.get('password', '').strip()
    confirm_password = request.form.get('confirm_password', '').strip()
    smtp_host = existing.get('smtp_host', '') or 'owa.fonafe.gob.pe'
    smtp_port = _parse_smtp_port(existing.get('smtp_port', '') or '587')
    smtp_security = _normalize_smtp_security(existing.get('smtp_security', '') or 'starttls')

    stored_sender = str(existing.get('sender', '')).strip()
    stored_password = str(existing.get('password', '')).strip()

    if not sender:
        session['system_authenticated'] = False
        session['smtp_authenticated'] = False
        session['smtp_link_verified'] = False
        if is_login_route:
            _register_login_failure('Ingresa tu correo para iniciar sesión con Microsoft 365.')
        else:
            session['login_message'] = 'Ingresa tu correo para iniciar sesión con Microsoft 365.'
        return redirect(url_for('iniciar_sesion'))

    if not password:
        same_saved_account = bool(stored_sender and stored_password and sender.lower() == stored_sender.lower())
        if same_saved_account:
            password = stored_password
            confirm_password = stored_password
        else:
            session['system_authenticated'] = False
            session['smtp_authenticated'] = False
            session['smtp_link_verified'] = False
            if is_login_route:
                _register_login_failure('No hay una contraseña guardada para ese correo. Configura la cuenta primero.')
            else:
                session['login_message'] = 'No hay una contraseña guardada para ese correo. Configura la cuenta primero.'
            return redirect(url_for('iniciar_sesion'))

    if not confirm_password:
        confirm_password = password

    if password != confirm_password:
        session['system_authenticated'] = False
        session['smtp_authenticated'] = False
        session['smtp_link_verified'] = False
        if is_login_route:
            _register_login_failure('Las contraseñas no coinciden. Verifica e intenta nuevamente.')
        else:
            session['login_message'] = 'Las contraseñas no coinciden. Verifica e intenta nuevamente.'
        return redirect(url_for('iniciar_sesion'))

    is_valid, validated_sender, used_security, error_message = _validate_office365_and_smtp_login(
        sender,
        password,
        smtp_host,
        smtp_port,
        smtp_security
    )
    if not is_valid:
        session['system_authenticated'] = False
        session['smtp_authenticated'] = False
        session['smtp_link_verified'] = False
        session['worker_login_via_graph'] = False
        session['worker_auth_method'] = ''
        if is_login_route:
            _register_login_failure(error_message)
        else:
            session['login_message'] = error_message
        return redirect(url_for('iniciar_sesion'))

    try:
        save_secure_smtp_credentials(
            validated_sender,
            password,
            smtp_host_value=smtp_host,
            smtp_port_value=str(smtp_port),
            smtp_security_value=used_security
        )
        sync_email_config_to_modules(validated_sender)
        session['system_authenticated'] = True
        session['smtp_authenticated'] = True
        session['smtp_link_verified'] = True
        session['worker_sender'] = validated_sender
        session['worker_password'] = password
        session['worker_smtp_host'] = smtp_host
        session['worker_smtp_port'] = str(smtp_port)
        session['worker_smtp_security'] = used_security
        session['worker_login_at'] = pd.Timestamp.now().strftime('%d/%m/%Y %H:%M:%S')
        session['worker_auth_method'] = str(session.get('worker_auth_method', '')).strip() or 'SMTP OWA + Microsoft 365 Graph'
        session['quick_password_verified'] = False
        if is_login_route:
            _reset_login_failures()
        add_account_activity('Correo electrónico', f'Sesión iniciada para {validated_sender}')
        session.pop('email_settings_message', None)
        session.pop('login_message', None)
    except Exception as ex:
        session['system_authenticated'] = False
        session['smtp_authenticated'] = False
        session['smtp_link_verified'] = False
        session['worker_login_via_graph'] = False
        session['worker_auth_method'] = ''
        if is_login_route:
            _register_login_failure(f'Error iniciando sesión: {ex}')
        else:
            session['login_message'] = f'Error iniciando sesión: {ex}'
        return redirect(url_for('iniciar_sesion'))

    return redirect(url_for('basedatos'))


@app.route('/correo_electronico/verificar_vinculo', methods=['POST'])
def correo_electronico_verificar_vinculo():
    auth_redirect = _require_worker_microsoft_login()
    if auth_redirect is not None:
        return auth_redirect

    sender = str(session.get('worker_sender', '')).strip()
    password = str(session.get('worker_password', '')).strip()
    smtp_host = str(session.get('worker_smtp_host', '')).strip() or 'owa.fonafe.gob.pe'
    smtp_port = _parse_smtp_port(session.get('worker_smtp_port', '') or '587')
    smtp_security = _normalize_smtp_security(session.get('worker_smtp_security', '') or 'starttls')

    if not sender or not password:
        session['smtp_authenticated'] = False
        session['smtp_link_verified'] = False
        session['system_authenticated'] = False
        session['login_message'] = 'No hay una sesión Microsoft activa para verificar. Inicia sesión nuevamente.'
        return redirect(url_for('iniciar_sesion'))

    is_valid, validated_sender, used_security, error_message = _validate_office365_and_smtp_login(
        sender,
        password,
        smtp_host,
        smtp_port,
        smtp_security
    )

    if not is_valid:
        session['smtp_authenticated'] = False
        session['smtp_link_verified'] = False
        session['system_authenticated'] = False
        session['worker_login_via_graph'] = False
        session['worker_auth_method'] = ''
        session['login_message'] = f'Vínculo no válido: {error_message}'
        return redirect(url_for('iniciar_sesion'))

    session['system_authenticated'] = True
    session['smtp_authenticated'] = True
    session['smtp_link_verified'] = True
    session['worker_sender'] = validated_sender
    session['worker_smtp_security'] = used_security
    session['worker_login_at'] = pd.Timestamp.now().strftime('%d/%m/%Y %H:%M:%S')
    session['worker_auth_method'] = str(session.get('worker_auth_method', '')).strip() or 'SMTP OWA'
    session['email_settings_message'] = '✅ Vínculo Microsoft 365 verificado correctamente.'
    add_account_activity('Correo electrónico', f'Vínculo verificado para {validated_sender}')

    return redirect(url_for('correo_electronico'))


@app.route('/configurar_correo', methods=['POST'])
def configurar_correo():
    return_to = str(request.form.get('return_to', '')).strip().lower()
    redirect_to_correos = return_to == 'correos'

    def _redirect_after_password(message_text, is_error=False):
        if redirect_to_correos:
            if is_error:
                session['quick_password_message'] = message_text
                session['quick_password_verified'] = False
                return redirect(url_for('correos', open_quick_password='1'))
            session['config_message'] = message_text
            return redirect(url_for('correos'))

        session['email_settings_message'] = message_text
        return redirect(url_for('correo_electronico'))

    auth_redirect = _require_worker_microsoft_login()
    if auth_redirect is not None:
        return auth_redirect

    existing = load_secure_smtp_credentials()
    sender = request.form.get('sender', '').strip() or existing.get('sender', '').strip()
    password = request.form.get('password', '').strip() or existing.get('password', '').strip()
    smtp_host = request.form.get('smtp_host', '').strip() or existing.get('smtp_host', '').strip() or 'owa.fonafe.gob.pe'
    smtp_port = _parse_smtp_port(request.form.get('smtp_port', '').strip() or existing.get('smtp_port', '').strip() or '587')
    smtp_security = _normalize_smtp_security(request.form.get('smtp_security', '').strip().lower() or existing.get('smtp_security', '').strip().lower() or 'starttls')

    if not sender or not password:
        return _redirect_after_password('Completa remitente y contraseña para continuar.', is_error=True)

    is_valid, validated_sender, used_security, error_message = _validate_office365_and_smtp_login(
        sender,
        password,
        smtp_host,
        smtp_port,
        smtp_security
    )
    if not is_valid:
        session['smtp_link_verified'] = False
        session['worker_login_via_graph'] = False
        session['worker_auth_method'] = ''
        return _redirect_after_password('Contraseña incorrecta. Intenta de nuevo.', is_error=True)

    try:
        save_secure_smtp_credentials(
            validated_sender,
            password,
            smtp_host_value=smtp_host,
            smtp_port_value=str(smtp_port),
            smtp_security_value=used_security
        )
        sync_email_config_to_modules(validated_sender)
        session['system_authenticated'] = True
        session['smtp_authenticated'] = True
        session['smtp_link_verified'] = True
        session['worker_sender'] = validated_sender
        session['worker_password'] = password
        session['worker_smtp_host'] = smtp_host
        session['worker_smtp_port'] = str(smtp_port)
        session['worker_smtp_security'] = used_security
        session['worker_auth_method'] = str(session.get('worker_auth_method', '')).strip() or 'SMTP OWA'
        session['quick_password_verified'] = True
        add_account_activity('Contraseña de correo', f'Sesión iniciada para {validated_sender}')
        return _redirect_after_password('✅ Contraseña de correo actualizada correctamente.')
    except Exception as ex:
        return _redirect_after_password('No se pudo guardar la contraseña de correo. Intenta nuevamente.', is_error=True)


@app.route('/configurar_google', methods=['POST'])
def configurar_google():
    existing = load_google_config()

    project_id = request.form.get('project_id', '').strip()
    client_id = request.form.get('client_id', '').strip()
    client_secret = request.form.get('client_secret', '').strip() or existing.get('client_secret', '')
    redirect_uri = request.form.get('redirect_uri', '').strip()
    scopes = request.form.get('scopes', '').strip() or 'openid email profile'
    api_key = request.form.get('api_key', '').strip() or existing.get('api_key', '')
    enabled = request.form.get('google_enabled', '').strip().lower() == 'on'

    if not client_id or not redirect_uri:
        session['config_message'] = 'Completa Client ID y Redirect URI para guardar Google.'
        return redirect(url_for('correos'))

    try:
        save_google_config({
            'project_id': project_id,
            'client_id': client_id,
            'client_secret': client_secret,
            'redirect_uri': redirect_uri,
            'scopes': scopes,
            'api_key': api_key,
            'enabled': enabled
        })
        session['config_message'] = '✅ Configuración Google guardada correctamente.'
    except Exception as ex:
        session['config_message'] = f'Error guardando configuración Google: {ex}'

    return redirect(url_for('correos'))


@app.route('/configuracion_inicio', methods=['POST'])
def configuracion_inicio():
    settings = load_general_settings()
    settings['inicio']['mostrar_resumen'] = request.form.get('mostrar_resumen', '').strip().lower() == 'on'
    settings['inicio']['notificaciones_activas'] = request.form.get('notificaciones_activas', '').strip().lower() == 'on'

    try:
        save_general_settings(settings)
        add_account_activity('Inicio', 'Preferencias de inicio actualizadas')
        session['config_message'] = '✅ Configuración de Inicio guardada correctamente.'
    except Exception as ex:
        session['config_message'] = f'Error guardando Inicio: {ex}'

    return redirect(url_for('correos'))


@app.route('/configuracion_perfil', methods=['POST'])
def configuracion_perfil():
    settings = load_general_settings()
    settings['perfil']['nombre_mostrar'] = request.form.get('nombre_mostrar', '').strip()
    settings['perfil']['cargo'] = request.form.get('cargo', '').strip()
    settings['perfil']['telefono'] = request.form.get('telefono', '').strip()
    settings['perfil']['genero'] = request.form.get('genero', settings['perfil'].get('genero', '')).strip()
    settings['perfil']['correo_personal'] = request.form.get('correo_personal', settings['perfil'].get('correo_personal', '')).strip()
    settings['perfil']['correo_alterno'] = request.form.get('correo_alterno', settings['perfil'].get('correo_alterno', '')).strip()
    settings['perfil']['fecha_nacimiento'] = request.form.get('fecha_nacimiento', settings['perfil'].get('fecha_nacimiento', '')).strip()
    settings['perfil']['idioma'] = request.form.get('idioma', settings['perfil'].get('idioma', 'Español (España)')).strip()
    settings['perfil']['direccion_casa'] = request.form.get('direccion_casa', settings['perfil'].get('direccion_casa', '')).strip()
    settings['perfil']['direccion_trabajo'] = request.form.get('direccion_trabajo', settings['perfil'].get('direccion_trabajo', '')).strip()
    settings['perfil']['otras_direcciones'] = request.form.get('otras_direcciones', settings['perfil'].get('otras_direcciones', '')).strip()

    try:
        save_general_settings(settings)
        add_account_activity('Perfil', f"Perfil actualizado: {settings['perfil']['nombre_mostrar']}")
        session['config_message'] = '✅ Configuración de Perfil guardada correctamente.'
    except Exception as ex:
        session['config_message'] = f'Error guardando Perfil: {ex}'

    return redirect(url_for('correos'))


@app.route('/configuracion_seguridad', methods=['POST'])
def configuracion_seguridad():
    settings = load_general_settings()
    payload = load_account_features()

    settings['seguridad']['doble_factor'] = request.form.get('doble_factor', '').strip().lower() == 'on'
    settings['seguridad']['cerrar_sesiones'] = request.form.get('cerrar_sesiones', '').strip().lower() == 'on'

    feature_security = payload.get('security', {}) if isinstance(payload.get('security', {}), dict) else {}

    feature_security['recovery_phone'] = request.form.get('recovery_phone', '').strip()
    feature_security['recovery_email'] = request.form.get('recovery_email', '').strip().lower()
    feature_security['two_factor_phone'] = request.form.get('two_factor_phone', '').strip()
    feature_security['backup_codes'] = max(_sanitize_int(request.form.get('backup_codes', '0'), 0), 0)
    feature_security['device_prompt_count'] = max(_sanitize_int(request.form.get('device_prompt_count', request.form.get('google_prompt_devices', '0')), 0), 0)
    feature_security['skip_password_when_possible'] = request.form.get('skip_password_when_possible', '').strip().lower() == 'on'
    feature_security['enhanced_browsing'] = request.form.get('enhanced_browsing', '').strip().lower() == 'on'
    payload['security'] = feature_security

    security_action = request.form.get('security_action', '').strip().lower()
    closed_sessions = 0
    if security_action == 'close_other_sessions':
        current_session_id = str(session.get('account_session_id', '')).strip()
        closed_sessions = _deactivate_other_sessions(payload, current_session_id)

    try:
        save_general_settings(settings)
        save_account_features(payload)
        if closed_sessions > 0:
            add_account_activity('Seguridad', f'Preferencias actualizadas y {closed_sessions} sesión(es) cerradas')
            session['config_message'] = f'✅ Seguridad guardada y {closed_sessions} sesión(es) cerradas en otros dispositivos.'
        elif security_action == 'close_other_sessions':
            add_account_activity('Seguridad', 'Preferencias actualizadas y sin sesiones adicionales para cerrar')
            session['config_message'] = '✅ Seguridad guardada. No había sesiones adicionales activas para cerrar.'
        else:
            add_account_activity('Seguridad', 'Preferencias de seguridad actualizadas')
            session['config_message'] = '✅ Configuración de Seguridad guardada correctamente.'
    except Exception as ex:
        session['config_message'] = f'Error guardando Seguridad: {ex}'

    return redirect(url_for('correos'))


@app.route('/configuracion_privacidad', methods=['POST'])
def configuracion_privacidad():
    settings = load_general_settings()
    payload = load_account_features()
    privacy_center = _ensure_privacy_center(payload)

    settings['privacidad']['compartir_datos'] = request.form.get('compartir_datos', '').strip().lower() == 'on'
    settings['privacidad']['analytics'] = request.form.get('analytics', '').strip().lower() == 'on'

    privacy_center['web_activity'] = settings['privacidad']['analytics']
    privacy_center['partner_ads'] = settings['privacidad']['compartir_datos']
    privacy_center['search_personalization'] = request.form.get('search_personalization', '').strip().lower() == 'on'
    privacy_center['play_personalization'] = request.form.get('play_personalization', '').strip().lower() == 'on'
    privacy_center['play_history'] = request.form.get('play_history', '').strip().lower() == 'on'
    privacy_center['youtube_history'] = request.form.get('youtube_history', '').strip().lower() == 'on'
    privacy_center['maps_timeline'] = request.form.get('maps_timeline', '').strip().lower() == 'on'
    privacy_center['ads_personalized'] = request.form.get('ads_personalized', '').strip().lower() == 'on'
    privacy_center['service_emails_enabled'] = request.form.get('service_emails_enabled', '').strip().lower() == 'on'
    privacy_center['fit_data_enabled'] = request.form.get('fit_data_enabled', '').strip().lower() == 'on'
    privacy_center['voice_match_enabled'] = request.form.get('voice_match_enabled', '').strip().lower() == 'on'

    payload['privacy_center'] = privacy_center

    try:
        save_general_settings(settings)
        save_account_features(payload)
        add_account_activity('Privacidad', 'Preferencias de privacidad actualizadas')
        session['config_message'] = '✅ Configuración de Privacidad guardada correctamente.'
    except Exception as ex:
        session['config_message'] = f'Error guardando Privacidad: {ex}'

    return redirect(url_for('correos'))


@app.route('/configuracion_contactos', methods=['POST'])
def configuracion_contactos():
    settings = load_general_settings()
    payload = load_account_features()

    correo_soporte = request.form.get('correo_soporte', '').strip().lower()
    correo_copia = request.form.get('correo_copia', '').strip().lower()

    settings['contactos']['correo_soporte'] = correo_soporte
    settings['contactos']['correo_copia'] = correo_copia

    contacts_share = _ensure_contacts_share(payload, settings)
    preferences = contacts_share.get('preferences', {})
    family = contacts_share.get('family', {})

    preferences['sync_interactions'] = request.form.get('sync_interactions', '').strip().lower() == 'on'
    preferences['sync_device_contacts'] = request.form.get('sync_device_contacts', '').strip().lower() == 'on'
    preferences['share_location'] = request.form.get('share_location', '').strip().lower() == 'on'
    preferences['profile_visible'] = request.form.get('profile_visible', '').strip().lower() == 'on'

    family['enabled'] = request.form.get('family_enabled', '').strip().lower() == 'on'
    family_name = request.form.get('group_name', '').strip()
    if family_name:
        family['group_name'] = family_name

    contacts_share['preferences'] = preferences
    contacts_share['family'] = family
    payload['contacts_share'] = contacts_share

    if correo_soporte:
        contacts = contacts_share.get('contacts', [])
        exists_support = False
        for item in contacts:
            if str(item.get('email', '')).strip().lower() == correo_soporte:
                exists_support = True
                if not str(item.get('name', '')).strip():
                    item['name'] = 'Contacto soporte'
                break

        if not exists_support:
            contacts.append({
                'id': _next_numeric_id(contacts),
                'name': 'Contacto soporte',
                'email': correo_soporte,
                'phone': '',
                'source': 'Configuración',
                'blocked': False
            })
        contacts_share['contacts'] = contacts
        payload['contacts_share'] = contacts_share

    try:
        save_general_settings(settings)
        save_account_features(payload)
        add_account_activity('Contactos', 'Contactos y compartir actualizado desde configuración')
        session['config_message'] = '✅ Configuración de Contactos guardada correctamente.'
    except Exception as ex:
        session['config_message'] = f'Error guardando Contactos: {ex}'

    return redirect(url_for('correos'))


@app.route('/contactos_compartir', methods=['GET'])
def contactos_compartir():
    add_account_activity('Contactos y compartir', 'Apertura de panel')

    settings = load_general_settings()
    payload = load_account_features()
    contacts_share = _ensure_contacts_share(payload, settings)
    save_account_features(payload)

    summary = _contacts_share_summary(contacts_share)
    message = session.pop('contacts_share_message', None)

    contacts_list = contacts_share.get('contacts', [])
    active_contacts = [item for item in contacts_list if not bool(item.get('blocked', False))]
    blocked_contacts = [item for item in contacts_list if bool(item.get('blocked', False))]

    return render_template(
        'contactos_compartir.html',
        message=message,
        settings=settings,
        contacts_share=contacts_share,
        contacts_summary=summary,
        active_contacts=active_contacts,
        blocked_contacts=blocked_contacts,
        profiles=contacts_share.get('profiles', []),
        connected_sender=get_connected_account_email()
    )


@app.route('/contactos_compartir/guardar', methods=['POST'])
def contactos_compartir_guardar():
    settings = load_general_settings()
    payload = load_account_features()
    contacts_share = _ensure_contacts_share(payload, settings)

    family = contacts_share.get('family', {})
    preferences = contacts_share.get('preferences', {})

    family['enabled'] = request.form.get('family_enabled', '').strip().lower() == 'on'
    family_name = request.form.get('group_name', '').strip()
    if family_name:
        family['group_name'] = family_name

    invites_remaining = _sanitize_int(request.form.get('invites_remaining', family.get('invites_remaining', 5)), family.get('invites_remaining', 5))
    family['invites_remaining'] = max(invites_remaining, 0)

    preferences['sync_interactions'] = request.form.get('sync_interactions', '').strip().lower() == 'on'
    preferences['sync_device_contacts'] = request.form.get('sync_device_contacts', '').strip().lower() == 'on'
    preferences['share_location'] = request.form.get('share_location', '').strip().lower() == 'on'
    preferences['profile_visible'] = request.form.get('profile_visible', '').strip().lower() == 'on'

    contacts_share['family'] = family
    contacts_share['preferences'] = preferences
    payload['contacts_share'] = contacts_share

    try:
        save_account_features(payload)
        add_account_activity('Contactos y compartir', 'Preferencias de contacto actualizadas')
        session['contacts_share_message'] = '✅ Preferencias de Contactos y compartir guardadas.'
    except Exception as ex:
        session['contacts_share_message'] = f'Error guardando preferencias: {ex}'

    return redirect(url_for('contactos_compartir'))


@app.route('/contactos_compartir/invitar_familia', methods=['POST'])
def contactos_compartir_invitar_familia():
    payload = load_account_features()
    contacts_share = _ensure_contacts_share(payload)
    family = contacts_share.get('family', {})

    invites_remaining = max(_sanitize_int(family.get('invites_remaining', 0), 0), 0)
    if invites_remaining <= 0:
        session['contacts_share_message'] = 'No quedan invitaciones disponibles para el grupo familiar.'
        return redirect(url_for('contactos_compartir'))

    family['invites_remaining'] = invites_remaining - 1
    contacts_share['family'] = family
    payload['contacts_share'] = contacts_share
    save_account_features(payload)

    add_account_activity('Contactos y compartir', 'Se envió una invitación familiar')
    session['contacts_share_message'] = '✅ Invitación familiar enviada correctamente.'
    return redirect(url_for('contactos_compartir'))


@app.route('/contactos_compartir/contacto/agregar', methods=['POST'])
def contactos_compartir_contacto_agregar():
    payload = load_account_features()
    contacts_share = _ensure_contacts_share(payload)
    contacts = contacts_share.get('contacts', [])

    name = request.form.get('name', '').strip()
    email = request.form.get('email', '').strip().lower()
    phone = request.form.get('phone', '').strip()
    source = request.form.get('source', '').strip() or 'Manual'

    if not name:
        session['contacts_share_message'] = 'Ingresa un nombre para el contacto.'
        return redirect(url_for('contactos_compartir'))

    if not email and not phone:
        session['contacts_share_message'] = 'Ingresa al menos correo o teléfono del contacto.'
        return redirect(url_for('contactos_compartir'))

    contacts.append({
        'id': _next_numeric_id(contacts),
        'name': name,
        'email': email,
        'phone': phone,
        'source': source,
        'blocked': False
    })

    contacts_share['contacts'] = contacts
    payload['contacts_share'] = contacts_share
    save_account_features(payload)

    add_account_activity('Contactos y compartir', f'Contacto agregado: {name}')
    session['contacts_share_message'] = '✅ Contacto agregado correctamente.'
    return redirect(url_for('contactos_compartir'))


@app.route('/contactos_compartir/contacto/eliminar/<int:contact_id>', methods=['POST'])
def contactos_compartir_contacto_eliminar(contact_id):
    payload = load_account_features()
    contacts_share = _ensure_contacts_share(payload)
    contacts = contacts_share.get('contacts', [])

    kept = [item for item in contacts if _sanitize_int(item.get('id', 0), 0) != int(contact_id)]
    if len(kept) == len(contacts):
        session['contacts_share_message'] = 'No se encontró el contacto seleccionado.'
        return redirect(url_for('contactos_compartir'))

    contacts_share['contacts'] = kept
    payload['contacts_share'] = contacts_share
    save_account_features(payload)

    add_account_activity('Contactos y compartir', f'Contacto eliminado: id={contact_id}')
    session['contacts_share_message'] = '✅ Contacto eliminado correctamente.'
    return redirect(url_for('contactos_compartir'))


@app.route('/contactos_compartir/contacto/bloquear/<int:contact_id>', methods=['POST'])
def contactos_compartir_contacto_bloquear(contact_id):
    payload = load_account_features()
    contacts_share = _ensure_contacts_share(payload)
    contacts = contacts_share.get('contacts', [])

    changed = False
    for item in contacts:
        if _sanitize_int(item.get('id', 0), 0) == int(contact_id):
            item['blocked'] = True
            changed = True
            break

    if not changed:
        session['contacts_share_message'] = 'No se encontró el contacto seleccionado.'
        return redirect(url_for('contactos_compartir'))

    contacts_share['contacts'] = contacts
    payload['contacts_share'] = contacts_share
    save_account_features(payload)

    add_account_activity('Contactos y compartir', f'Contacto bloqueado: id={contact_id}')
    session['contacts_share_message'] = '✅ Contacto bloqueado correctamente.'
    return redirect(url_for('contactos_compartir'))


@app.route('/contactos_compartir/contacto/desbloquear/<int:contact_id>', methods=['POST'])
def contactos_compartir_contacto_desbloquear(contact_id):
    payload = load_account_features()
    contacts_share = _ensure_contacts_share(payload)
    contacts = contacts_share.get('contacts', [])

    changed = False
    for item in contacts:
        if _sanitize_int(item.get('id', 0), 0) == int(contact_id):
            item['blocked'] = False
            changed = True
            break

    if not changed:
        session['contacts_share_message'] = 'No se encontró el contacto seleccionado.'
        return redirect(url_for('contactos_compartir'))

    contacts_share['contacts'] = contacts
    payload['contacts_share'] = contacts_share
    save_account_features(payload)

    add_account_activity('Contactos y compartir', f'Contacto desbloqueado: id={contact_id}')
    session['contacts_share_message'] = '✅ Contacto desbloqueado correctamente.'
    return redirect(url_for('contactos_compartir'))


@app.route('/contactos_compartir/perfil/agregar', methods=['POST'])
def contactos_compartir_perfil_agregar():
    payload = load_account_features()
    contacts_share = _ensure_contacts_share(payload)
    profiles = contacts_share.get('profiles', [])

    profile_name = request.form.get('profile_name', '').strip()
    profile_description = request.form.get('profile_description', '').strip() or 'Perfil para servicios conectados'

    if not profile_name:
        session['contacts_share_message'] = 'Ingresa el nombre del perfil.'
        return redirect(url_for('contactos_compartir'))

    profiles.append({
        'id': _next_numeric_id(profiles),
        'name': profile_name,
        'description': profile_description
    })

    contacts_share['profiles'] = profiles
    payload['contacts_share'] = contacts_share
    save_account_features(payload)

    add_account_activity('Contactos y compartir', f'Perfil agregado: {profile_name}')
    session['contacts_share_message'] = '✅ Perfil agregado correctamente.'
    return redirect(url_for('contactos_compartir'))


@app.route('/subir_foto_perfil', methods=['POST'])
def subir_foto_perfil():
    image_file = request.files.get('profile_image')
    next_url = request.form.get('next', '').strip()
    if image_file is None or image_file.filename == '':
        session['config_message'] = 'Selecciona una imagen para subir.'
        if next_url == 'informacion_personal':
            return redirect(url_for('informacion_personal'))
        if next_url == 'home':
            return redirect(url_for('home'))
        return redirect(url_for('correos'))

    if not _allowed_image_extension(image_file.filename):
        session['config_message'] = 'Formato no válido. Usa PNG, JPG, JPEG o WEBP.'
        if next_url == 'informacion_personal':
            return redirect(url_for('informacion_personal'))
        if next_url == 'home':
            return redirect(url_for('home'))
        return redirect(url_for('correos'))

    try:
        os.makedirs(_profile_photo_dir(), exist_ok=True)
        photo_path = _profile_photo_path()
        image_file.save(photo_path)

        settings = load_general_settings()
        current_version = int(settings.get('perfil', {}).get('foto_version', 0) or 0)
        settings['perfil']['foto_version'] = current_version + 1
        save_general_settings(settings)
        add_account_activity('Perfil', 'Foto de perfil actualizada')

        session['config_message'] = '✅ Foto de perfil actualizada correctamente.'
    except Exception as ex:
        session['config_message'] = f'Error subiendo foto de perfil: {ex}'

    if next_url == 'informacion_personal':
        if 'config_message' in session:
            session['personal_info_message'] = session.pop('config_message')
        return redirect(url_for('informacion_personal'))

    if next_url == 'home':
        return redirect(url_for('home'))

    return redirect(url_for('correos'))


@app.route('/foto_perfil_actual', methods=['GET'])
def foto_perfil_actual():
    photo_path = _profile_photo_path()
    if not os.path.exists(photo_path):
        return '', 404
    return send_file(photo_path)

@app.route('/upload', methods=['POST','GET'])
def upload():
    
    if request.method == 'POST':
        ruta_archivo = files_path('movimientos.xlsx')
        return send_file(ruta_archivo, as_attachment=True, download_name="movimientos.xlsx")
   
    

@app.route('/download_recaudos', methods=['POST'])
def download_recaudos():
    ruta_archivo = files_path('recaudos.xlsx')
    return send_file(ruta_archivo, as_attachment=True, download_name="recaudos.xlsx")

def guardaAsientos(movimientosAsientos):
    movimientosAsientos["Fecha"] = pd.to_datetime(movimientosAsientos["Fecha"], format="%d/%m/%Y").dt.strftime("%d/%m/%Y")
    excel_file = BytesIO()
    movimientosAsientos.to_excel(excel_file, index=False)
    excel_file.seek(0)
    workbook = openpyxl.load_workbook(excel_file)
    worksheet = workbook.active 
    worksheet.column_dimensions["A"].width = 20  
    worksheet.column_dimensions["C"].width = 30  
    worksheet.column_dimensions["K"].width = 40  
    worksheet.column_dimensions["L"].width = 40  
    worksheet.column_dimensions["M"].width = 35 
    worksheet.column_dimensions["N"].width = 28 

    ruta_archivo = files_path('asientos.xlsx')
    workbook.save(ruta_archivo)


@app.route('/download_asientos', methods=['POST'])
def dowload_asientos():
    ruta_archivo = files_path('asientos.xlsx')
    return send_file(ruta_archivo, as_attachment=True, download_name="Asiento.xlsx")


@app.route('/send_emails', methods=['POST'])
def send_emails():
    auth_redirect = _require_worker_microsoft_login()
    if auth_redirect is not None:
        return auth_redirect

    def _redirect_correos_with_message(message_text):
        session['config_message'] = message_text
        return redirect(url_for('correos'))

    emails = session.get('asiento_emails', [])
    if not emails:
        return _redirect_correos_with_message('No hay correos para enviar')

    manual_confirm = request.form.get('manual_confirm', '').strip().lower()
    if manual_confirm != 'yes':
        return _redirect_correos_with_message('Marca la confirmación de envío manual antes de enviar correos.')

    selected_emails = request.form.getlist('selected_emails')
    if not selected_emails:
        return _redirect_correos_with_message('Selecciona al menos un correo para enviar. No se envió nada automáticamente.')

    emails_to_send = sorted(set(selected_emails))
    
    # Obtener vouchers generados de la sesión
    vouchers_generados = session.get('vouchers_generados', [])
    
    # Log de diagnóstico
    logging.info(f'Total vouchers en sesión: {len(vouchers_generados)}')
    for v in vouchers_generados:
        logging.info(f"Voucher disponible: email={v.get('email')}, ref={v.get('referencia')}, path={v.get('filepath')}")
    
    # Crear diccionario para buscar voucher por email
    vouchers_por_email = {}
    for voucher in vouchers_generados:
        email = voucher.get('email', '').strip().lower()
        if email:
            vouchers_por_email[email] = voucher
            logging.info(f"Voucher indexado para: {email}")
    
    # Configuración para Microsoft 365 (Outlook)
    secure_smtp = load_secure_smtp_credentials()
    form_sender = request.form.get('sender', '').strip()
    form_password = request.form.get('password', '').strip()
    form_smtp_host = request.form.get('smtp_host', '').strip()
    form_smtp_port = request.form.get('smtp_port', '').strip()
    form_smtp_security = request.form.get('smtp_security', '').strip().lower()

    session_sender = str(session.get('worker_sender', '')).strip()
    session_password = str(session.get('worker_password', '')).strip()
    session_smtp_host = str(session.get('worker_smtp_host', '')).strip()
    session_smtp_port = str(session.get('worker_smtp_port', '')).strip()
    session_smtp_security = str(session.get('worker_smtp_security', '')).strip().lower()

    sender = form_sender or session_sender or os.environ.get('OUTLOOK_SENDER', '').strip() or secure_smtp.get('sender', '').strip()
    password = form_password or session_password or os.environ.get('OUTLOOK_PASSWORD', '').strip() or secure_smtp.get('password', '').strip()
    subject_env = os.environ.get('OUTLOOK_SUBJECT', '').strip()
    subject = subject_env or 'Confirmación de Abono Recibido - ENSA'
    subject = re.sub(r'distriluz\s+ensa', 'ENSA', subject, flags=re.IGNORECASE).strip()
    smtp_host = form_smtp_host or session_smtp_host or os.environ.get('OUTLOOK_SMTP_HOST', '').strip() or secure_smtp.get('smtp_host', '').strip() or 'owa.fonafe.gob.pe'
    smtp_port_raw = form_smtp_port or session_smtp_port or os.environ.get('OUTLOOK_SMTP_PORT', '').strip() or secure_smtp.get('smtp_port', '').strip() or '587'
    smtp_security = (form_smtp_security or session_smtp_security or os.environ.get('OUTLOOK_SMTP_SECURITY', '').strip() or secure_smtp.get('smtp_security', '').strip() or 'starttls').lower()
    try:
        smtp_port = int(smtp_port_raw)
    except ValueError:
        smtp_port = 587
        logging.warning(f"OUTLOOK_SMTP_PORT inválido ('{smtp_port_raw}'). Usando 587 por defecto.")

    if smtp_security not in ('ssl', 'starttls', 'auto'):
        logging.warning(f"OUTLOOK_SMTP_SECURITY inválido ('{smtp_security}'). Usando 'starttls'.")
        smtp_security = 'starttls'

    should_save_smtp = request.form.get('save_smtp', 'yes').strip().lower() in ('yes', 'on', 'true', '1')
    if should_save_smtp and sender and password:
        try:
            save_secure_smtp_credentials(
                sender,
                password,
                smtp_host_value=smtp_host,
                smtp_port_value=str(smtp_port),
                smtp_security_value=smtp_security
            )
            sync_email_config_to_modules(sender)
        except Exception as save_ex:
            logging.warning(f'No se pudo persistir configuración SMTP desde envío manual: {save_ex}')

    if not sender or not password:
        return _redirect_correos_with_message(
            'Falta configurar correo remitente y clave SMTP. Ingresa Usuario y Contraseña en el panel de envío y vuelve a intentar.'
        )

    sent_ok = []
    sent_fail = []
    sent_with_voucher_html = []
    used_security = smtp_security
    used_port = smtp_port
    
    try:
        # Conexión SMTP configurable (ssl | starttls | auto)
        smtp_conn = None
        try:
            if smtp_security == 'ssl':
                smtp_conn = smtplib.SMTP_SSL(smtp_host, smtp_port, timeout=30)
                smtp_conn.ehlo()
                used_security = 'ssl'
                used_port = smtp_port
            else:
                smtp_conn = smtplib.SMTP(smtp_host, smtp_port, timeout=30)
                smtp_conn.ehlo()
                smtp_conn.starttls()
                smtp_conn.ehlo()
                used_security = 'starttls'
                used_port = smtp_port
        except Exception as conn_ex:
            wrong_version = 'WRONG_VERSION_NUMBER' in str(conn_ex).upper()
            can_retry_starttls = smtp_security in ('ssl', 'auto') and wrong_version
            if can_retry_starttls:
                logging.warning(
                    f"SMTP SSL no compatible en {smtp_host}:{smtp_port} ({conn_ex}). Reintentando con STARTTLS."
                )
                smtp_conn = smtplib.SMTP(smtp_host, smtp_port, timeout=30)
                smtp_conn.ehlo()
                smtp_conn.starttls()
                smtp_conn.ehlo()
                used_security = 'starttls'
                used_port = smtp_port
            elif _should_retry_with_port_25(smtp_port, conn_ex):
                logging.warning(
                    f"SMTP {smtp_host}:{smtp_port} rechazó conexión ({conn_ex}). Reintentando por puerto 25 con STARTTLS."
                )
                smtp_conn = smtplib.SMTP(smtp_host, 25, timeout=30)
                smtp_conn.ehlo()
                smtp_conn.starttls()
                smtp_conn.ehlo()
                used_security = 'starttls'
                used_port = 25
            else:
                raise

        with smtp_conn as smtp:
            try:
                smtp.login(sender, password)
            except smtplib.SMTPAuthenticationError as auth_ex:
                corrected_sender, sender_corrected = normalize_sender_email(sender)
                if sender_corrected and corrected_sender != sender:
                    try:
                        smtp.login(corrected_sender, password)
                        sender = corrected_sender
                        logging.warning(
                            'Se corrigió remitente con dominio typo para autenticación SMTP: %s',
                            sender
                        )
                    except smtplib.SMTPAuthenticationError:
                        return _redirect_correos_with_message(
                            'No se pudieron validar tus credenciales de correo. '
                            'El usuario guardado tiene dominio typo. Usa tu correo con @distriluz.com.pe en las credenciales SMTP. '
                            'No se envió ningún correo.'
                        )
                else:
                    return _redirect_correos_with_message(
                        'No se pudieron validar tus credenciales de correo. '
                        'Verifica usuario/clave en credenciales SMTP o confirma con TI que la cuenta tenga SMTP AUTH habilitado en owa.fonafe.gob.pe. '
                        'No se envió ningún correo.'
                    )

            if should_save_smtp and sender and password and (used_port != smtp_port or used_security != smtp_security):
                try:
                    save_secure_smtp_credentials(
                        sender,
                        password,
                        smtp_host_value=smtp_host,
                        smtp_port_value=str(used_port),
                        smtp_security_value=used_security
                    )
                    session['worker_smtp_port'] = str(used_port)
                    session['worker_smtp_security'] = used_security
                except Exception as save_ex:
                    logging.warning(f'No se pudo persistir fallback SMTP efectivo: {save_ex}')

            for recipient in emails_to_send:
                try:
                    recipient_lower = recipient.strip().lower()
                    voucher_info = vouchers_por_email.get(recipient_lower)

                    nombre_cliente = ''
                    if voucher_info:
                        nombre_cliente = str(voucher_info.get('nombre_cliente', '')).strip()

                    if not nombre_cliente:
                        local_part = recipient.split('@')[0]
                        name_tokens = [token for token in re.split(r'[._\-]+', local_part) if token and not token.isdigit()]
                        if name_tokens:
                            nombre_cliente = ' '.join(token.capitalize() for token in name_tokens[:3])

                    if not nombre_cliente:
                        nombre_cliente = 'Cliente'

                    saludo = build_saludo_cliente(nombre_cliente)
                    body_text = build_voucher_email_text(saludo)
                    body_html = build_voucher_email_html(saludo, voucher_info)

                    msg = EmailMessage()
                    msg['Subject'] = subject
                    msg['From'] = sender
                    msg['To'] = recipient
                    msg.set_content(body_text)
                    msg.add_alternative(body_html, subtype='html')

                    logging.info(f"Procesando email HTML: {recipient}")
                    
                    smtp.send_message(msg)

                    sent_ok.append(recipient)
                    sent_with_voucher_html.append(recipient)
                    logging.info(f"📧 Email enviado con voucher HTML a: {recipient}")
                        
                except Exception as send_ex:
                    logging.error(f"Error enviando a {recipient}: {send_ex}")
                    sent_fail.append(f"{recipient}: {send_ex}")

        report_path = files_path('emails_send_report.csv')
        with open(report_path, 'w', encoding='utf-8') as report:
            report.write('estado,email,detalle\n')
            for ok in sent_ok:
                report.write(f'sent,{ok},ok\n')
            for fail in sent_fail:
                email_part = fail.split(':', 1)[0]
                detail_part = fail.split(':', 1)[1].strip() if ':' in fail else 'error'
                report.write(f'failed,{email_part},"{detail_part}"\n')

        sent_ok_set = {email.strip().lower() for email in sent_ok}
        remaining_emails = [email for email in emails if email.strip().lower() not in sent_ok_set]
        emails_no_enviados = remaining_emails[:]

        if len(sent_ok) > 0:
            session['asiento_emails'] = remaining_emails
            save_emails_cache(remaining_emails)

            vouchers_restantes = []
            for voucher in vouchers_generados:
                voucher_email = str(voucher.get('email', '')).strip().lower()
                if voucher_email not in sent_ok_set:
                    vouchers_restantes.append(voucher)
            session['vouchers_generados'] = vouchers_restantes

            if remaining_emails:
                session.pop('asiento_email_warning', None)
            else:
                session.pop('asiento_email_warning', None)
        
        if len(sent_ok) == 0 and len(sent_fail) > 0:
            message = f"⚠️ Envío finalizado con errores. Enviados: {len(sent_ok)}. Fallidos: {len(sent_fail)}."
        elif len(sent_ok) > 0 and len(sent_fail) == 0:
            message = "✅ Envío exitosamente"
        else:
            message = f"✅ Envío finalizado. Enviados: {len(sent_ok)}. Fallidos: {len(sent_fail)}."
        
        # Información de vouchers HTML enviados
        if len(sent_with_voucher_html) > 0:
            message += f"\n📄 Con voucher HTML en el correo: {len(sent_with_voucher_html)}"
        
        if len(emails_no_enviados) > 0:
            message += f"\n\n⚠️ CORREOS NO ENVIADOS ({len(emails_no_enviados)}):\n"
            for email_no_enviado in emails_no_enviados:  # Mostrar TODOS sin límite
                message += f"• {email_no_enviado}\n"
        
        if len(sent_fail) > 0:
            message += '\n\nRevisa emails_send_report.csv para detalles de errores.'

        if len(sent_ok) > 0:
            message += '\n\nSi no lo ves en Bandeja de Entrada, revisa la carpeta Enviados del remitente.'

        return _redirect_correos_with_message(message)
    except Exception as ex:
        return _redirect_correos_with_message(
            f'Error enviando por Outlook/Microsoft 365 ({smtp_host}:{used_port}, {used_security}): {ex}'
        )


@app.route('/cerrar_sesion', methods=['GET'])
def cerrar_sesion():
    session.clear()
    return redirect(url_for('iniciar_sesion'))


@app.route('/vouchers')
def listar_vouchers():
    """
    Muestra la lista de vouchers generados en la sesión actual
    """
    vouchers = session.get('vouchers_generados', [])
    return render_template('vouchers.html', vouchers=vouchers)


@app.route('/voucher/<path:filename>')
def descargar_voucher(filename):
    """
    Descarga un voucher específico
    """
    try:
        filepath = vouchers_path(filename)
        if os.path.exists(filepath):
            return send_file(filepath, as_attachment=True, download_name=filename)
        else:
            return "Voucher no encontrado", 404
    except Exception as e:
        return f"Error descargando voucher: {str(e)}", 500


@app.route('/vouchers/descargar_todos')
def descargar_todos_vouchers():
    """
    Descarga todos los vouchers generados en un archivo ZIP
    """
    import zipfile
    from io import BytesIO
    
    vouchers = session.get('vouchers_generados', [])
    if not vouchers:
        return "No hay vouchers para descargar", 404
    
    try:
        # Crear archivo ZIP en memoria
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for voucher in vouchers:
                filepath = voucher['filepath']
                if os.path.exists(filepath):
                    filename = os.path.basename(filepath)
                    zip_file.write(filepath, filename)
        
        zip_buffer.seek(0)
        return send_file(
            zip_buffer,
            mimetype='application/zip',
            as_attachment=True,
            download_name='vouchers_asiento.zip'
        )
    except Exception as e:
        return f"Error creando archivo ZIP: {str(e)}", 500



@app.errorhandler(Exception)
def handle_exception(e):
     # Log full traceback to file
     tb = traceback.format_exc()
     logging.error('Unhandled exception:\n%s', tb)
     # return a friendly error page
     return render_template('error.html', message=str(e)), 500


if __name__ == '__main__':
    # Configuración para red local
    app.run(host='0.0.0.0', port=5000, debug=True)

    